"""GSM folder scanner: auto-detect billing and identification files.

Handles:
- ZIP archives (nested ZIPs too)
- XLSX billing files (T-Mobile, Orange, Play, Plus)
- XLSX/CSV identification files (Orange, Play, Plus)
- Skips irrelevant files (images, PDFs, docs, etc.)
"""

from __future__ import annotations

import logging
import shutil
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

log = logging.getLogger(__name__)

# Extensions we try to classify
_CANDIDATE_EXTS = {".xlsx", ".xls", ".csv", ".txt"}

# Extensions we silently skip
_SKIP_EXTS = {
    ".pdf", ".doc", ".docx", ".pptx", ".jpg", ".jpeg", ".png", ".gif",
    ".bmp", ".tiff", ".mp3", ".mp4", ".avi", ".mov", ".wav", ".ogg",
    ".html", ".htm", ".xml", ".json", ".ini", ".cfg", ".log", ".bat",
    ".exe", ".dll", ".sys", ".tmp", ".bak",
}


@dataclass
class ScannedFile:
    """Result of classifying one file."""

    filename: str
    path: Path
    file_type: str          # "billing", "identification", "unknown", "skipped"
    operator: str = ""      # e.g. "T-Mobile", "Orange", "Play", "Plus"
    operator_id: str = ""   # e.g. "tmobile", "orange", "play", "plus"
    confidence: float = 0.0
    detail: str = ""        # human-readable note
    size_bytes: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "file_type": self.file_type,
            "operator": self.operator,
            "operator_id": self.operator_id,
            "confidence": round(self.confidence, 2),
            "detail": self.detail,
            "size_bytes": self.size_bytes,
        }


@dataclass
class ScanResult:
    """Result of scanning a folder or set of files."""

    files: List[ScannedFile] = field(default_factory=list)
    billing_files: List[ScannedFile] = field(default_factory=list)
    identification_files: List[ScannedFile] = field(default_factory=list)
    skipped_files: List[ScannedFile] = field(default_factory=list)
    unknown_files: List[ScannedFile] = field(default_factory=list)
    zips_extracted: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "total_files": len(self.files),
            "billing_count": len(self.billing_files),
            "identification_count": len(self.identification_files),
            "skipped_count": len(self.skipped_files),
            "unknown_count": len(self.unknown_files),
            "zips_extracted": self.zips_extracted,
            "files": [f.to_dict() for f in self.files],
        }


def extract_zips(folder: Path, max_depth: int = 3) -> int:
    """Recursively extract ZIP archives in a folder. Returns count extracted."""
    if max_depth <= 0:
        return 0

    extracted = 0
    for zp in list(folder.rglob("*.zip")):
        try:
            target = zp.parent / zp.stem
            target.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(zp, "r") as zf:
                zf.extractall(target)
            extracted += 1
            log.info("Extracted ZIP: %s → %s", zp.name, target)
            # Remove the ZIP to avoid re-processing
            zp.unlink(missing_ok=True)
            # Recurse for nested ZIPs
            extracted += extract_zips(target, max_depth - 1)
        except (zipfile.BadZipFile, Exception) as e:
            log.warning("Cannot extract ZIP %s: %s", zp.name, e)

    return extracted


def _classify_xlsx(file_path: Path) -> ScannedFile:
    """Try to classify an XLSX file as billing or identification."""
    filename = file_path.name
    size = file_path.stat().st_size

    # --- Try as billing first ---
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        sheet_names = [s.lower() for s in wb.sheetnames]

        # Get first header row
        headers: List[str] = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(max_row=20, values_only=True):
                cells = [str(c).strip().lower() for c in row if c is not None]
                non_empty = [c for c in cells if c]
                if len(non_empty) >= 3:
                    headers = cells
                    break
            if headers:
                break
        wb.close()

        if headers:
            from .parsers.registry import detect_operator
            parser_cls, confidence = detect_operator(headers, sheet_names)
            if parser_cls is not None and confidence >= 0.3:
                return ScannedFile(
                    filename=filename,
                    path=file_path,
                    file_type="billing",
                    operator=getattr(parser_cls, "OPERATOR_NAME", ""),
                    operator_id=getattr(parser_cls, "OPERATOR_ID", ""),
                    confidence=confidence,
                    detail=f"Biling {getattr(parser_cls, 'OPERATOR_NAME', '?')}",
                    size_bytes=size,
                )
    except Exception as e:
        log.debug("XLSX billing detection error for %s: %s", filename, e)

    # --- Try as identification ---
    try:
        from .identification import detect_id_format
        id_fmt = detect_id_format(file_path)
        if id_fmt:
            op_names = {"orange": "Orange", "play": "Play", "plus": "Plus"}
            return ScannedFile(
                filename=filename,
                path=file_path,
                file_type="identification",
                operator=op_names.get(id_fmt, id_fmt),
                operator_id=id_fmt,
                confidence=0.8,
                detail=f"Identyfikacja {op_names.get(id_fmt, id_fmt)}",
                size_bytes=size,
            )
    except Exception as e:
        log.debug("XLSX identification detection error for %s: %s", filename, e)

    return ScannedFile(
        filename=filename,
        path=file_path,
        file_type="unknown",
        detail="Nie rozpoznano formatu XLSX",
        size_bytes=size,
    )


def _classify_csv(file_path: Path) -> ScannedFile:
    """Try to classify a CSV/TXT file as billing or identification."""
    filename = file_path.name
    size = file_path.stat().st_size

    # --- Try as Plus CSV billing first (custom quoting format) ---
    try:
        from .parsers.plus import is_plus_csv
        if is_plus_csv(file_path):
            return ScannedFile(
                filename=filename,
                path=file_path,
                file_type="billing",
                operator="Plus (Polkomtel)",
                operator_id="plus",
                confidence=0.95,
                detail="Biling Plus (Polkomtel) — CSV",
                size_bytes=size,
            )
    except Exception as e:
        log.debug("CSV billing detection error for %s: %s", filename, e)

    # --- Try as Play CSV billing ---
    try:
        from .parsers.play import is_play_csv
        if is_play_csv(file_path):
            return ScannedFile(
                filename=filename,
                path=file_path,
                file_type="billing",
                operator="Play (P4)",
                operator_id="play",
                confidence=0.95,
                detail="Biling Play (P4) — CSV",
                size_bytes=size,
            )
    except Exception as e:
        log.debug("CSV billing detection error for %s: %s", filename, e)

    # --- Try as identification ---
    try:
        from .identification import detect_id_format
        id_fmt = detect_id_format(file_path)
        if id_fmt:
            op_names = {"orange": "Orange", "play": "Play", "plus": "Plus"}
            return ScannedFile(
                filename=filename,
                path=file_path,
                file_type="identification",
                operator=op_names.get(id_fmt, id_fmt),
                operator_id=id_fmt,
                confidence=0.8,
                detail=f"Identyfikacja {op_names.get(id_fmt, id_fmt)}",
                size_bytes=size,
            )
    except Exception as e:
        log.debug("CSV identification detection error for %s: %s", filename, e)

    return ScannedFile(
        filename=filename,
        path=file_path,
        file_type="unknown",
        detail="Nie rozpoznano formatu CSV",
        size_bytes=size,
    )


def scan_folder(folder: Path) -> ScanResult:
    """Scan a folder for billing and identification files.

    Steps:
    1. Extract all ZIP archives (recursively)
    2. Walk all files, classify each as billing/identification/unknown/skipped
    3. Return categorised results
    """
    result = ScanResult()

    # Step 1: extract ZIPs
    result.zips_extracted = extract_zips(folder)

    # Step 2: classify all files
    for fp in sorted(folder.rglob("*")):
        if not fp.is_file():
            continue

        suffix = fp.suffix.lower()

        # Skip known irrelevant extensions
        if suffix in _SKIP_EXTS or suffix == ".zip":
            sf = ScannedFile(
                filename=fp.name,
                path=fp,
                file_type="skipped",
                detail=f"Pominięto ({suffix})",
                size_bytes=fp.stat().st_size,
            )
            result.files.append(sf)
            result.skipped_files.append(sf)
            continue

        # Not a candidate extension either — skip
        if suffix not in _CANDIDATE_EXTS:
            sf = ScannedFile(
                filename=fp.name,
                path=fp,
                file_type="skipped",
                detail=f"Nieobsługiwane rozszerzenie ({suffix})",
                size_bytes=fp.stat().st_size,
            )
            result.files.append(sf)
            result.skipped_files.append(sf)
            continue

        # Classify
        if suffix in (".xlsx", ".xls"):
            sf = _classify_xlsx(fp)
        else:
            sf = _classify_csv(fp)

        result.files.append(sf)

        if sf.file_type == "billing":
            result.billing_files.append(sf)
        elif sf.file_type == "identification":
            result.identification_files.append(sf)
        elif sf.file_type == "unknown":
            result.unknown_files.append(sf)
        else:
            result.skipped_files.append(sf)

    log.info(
        "Scan complete: %d files — %d billing, %d identification, %d unknown, %d skipped",
        len(result.files),
        len(result.billing_files),
        len(result.identification_files),
        len(result.unknown_files),
        len(result.skipped_files),
    )

    return result


def scan_files(file_paths: List[Path]) -> ScanResult:
    """Scan a list of individual files (no ZIP extraction, no folder walk).

    Useful when files are uploaded directly (not from a folder).
    """
    result = ScanResult()

    # If any are ZIPs, extract them to a temp folder first
    tmp_dirs: List[Path] = []
    non_zip_files: List[Path] = list(file_paths)
    zip_files = [fp for fp in non_zip_files if fp.suffix.lower() == ".zip"]

    for zp in zip_files:
        non_zip_files.remove(zp)
        tmp_dir = Path(tempfile.mkdtemp(prefix="gsm_scan_"))
        tmp_dirs.append(tmp_dir)
        try:
            target = tmp_dir / zp.stem
            target.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(zp, "r") as zf:
                zf.extractall(target)
            result.zips_extracted += 1
            # Recurse into extracted folder
            sub_result = scan_folder(target)
            result.files.extend(sub_result.files)
            result.billing_files.extend(sub_result.billing_files)
            result.identification_files.extend(sub_result.identification_files)
            result.skipped_files.extend(sub_result.skipped_files)
            result.unknown_files.extend(sub_result.unknown_files)
            result.zips_extracted += sub_result.zips_extracted
        except (zipfile.BadZipFile, Exception) as e:
            log.warning("Cannot extract ZIP %s: %s", zp.name, e)
            sf = ScannedFile(
                filename=zp.name, path=zp, file_type="skipped",
                detail=f"Błąd ZIP: {e}", size_bytes=zp.stat().st_size,
            )
            result.files.append(sf)
            result.skipped_files.append(sf)

    # Classify remaining non-ZIP files
    for fp in non_zip_files:
        suffix = fp.suffix.lower()

        if suffix in _SKIP_EXTS:
            sf = ScannedFile(
                filename=fp.name, path=fp, file_type="skipped",
                detail=f"Pominięto ({suffix})", size_bytes=fp.stat().st_size,
            )
            result.files.append(sf)
            result.skipped_files.append(sf)
            continue

        if suffix not in _CANDIDATE_EXTS:
            sf = ScannedFile(
                filename=fp.name, path=fp, file_type="skipped",
                detail=f"Nieobsługiwane rozszerzenie ({suffix})",
                size_bytes=fp.stat().st_size,
            )
            result.files.append(sf)
            result.skipped_files.append(sf)
            continue

        if suffix in (".xlsx", ".xls"):
            sf = _classify_xlsx(fp)
        else:
            sf = _classify_csv(fp)

        result.files.append(sf)

        if sf.file_type == "billing":
            result.billing_files.append(sf)
        elif sf.file_type == "identification":
            result.identification_files.append(sf)
        elif sf.file_type == "unknown":
            result.unknown_files.append(sf)
        else:
            result.skipped_files.append(sf)

    return result, tmp_dirs
