"""BTS (Base Transceiver Station) database management.

Provides SQLite-based storage for BTS/cell tower data from:
- UKE (Polish telecom regulator) CSV — has azimuth, antenna height, coordinates
- OpenCelliD CSV — has MCC/MNC/LAC/CID + crowdsourced coordinates + range
- Billing records (T-Mobile includes BTS X/Y/CI/LAC/Azimuth in CDRs)

Schema:
    bts_stations:
        mcc, mnc, lac, cid          — cell identity (primary key composite)
        lat, lon                     — WGS84 coordinates
        azimuth                      — antenna beam direction (degrees from north)
        range_m                      — estimated coverage radius (meters)
        radio                        — technology (GSM, UMTS, LTE, NR)
        city, street, address        — address info
        source                       — 'uke', 'opencellid', 'billing', 'manual'
        samples                      — number of measurements (OpenCelliD)
        antenna_height               — antenna height above ground (UKE)
        tilt                         — antenna tilt angle (UKE)
        updated                      — last update timestamp
"""

from __future__ import annotations

import csv
import io
import logging
import re
import sqlite3
import time
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

log = logging.getLogger(__name__)

# Poland MCC
MCC_POLAND = 260

# Polish MNC → operator name
MNC_OPERATORS: Dict[int, str] = {
    1: "Plus (Polkomtel)",
    2: "T-Mobile Polska",
    3: "Orange Polska",
    6: "Play (P4)",
    7: "Netia",
    9: "Lycamobile",
    10: "Sferia",
    11: "Nordisk Polska",
    12: "Cyfrowy Polsat",
    15: "Aero2",
    16: "Mobyland",
    17: "Aero2",
    34: "NetWorkS!",
    35: "Emitel",
    98: "Play (P4)",  # MVNO range
}


@dataclass
class BTSStation:
    """Single BTS station record."""

    mcc: int = MCC_POLAND
    mnc: int = 0
    lac: int = 0
    cid: int = 0
    lat: float = 0.0
    lon: float = 0.0
    azimuth: Optional[float] = None
    range_m: Optional[int] = None
    radio: str = ""
    city: str = ""
    street: str = ""
    address: str = ""
    source: str = ""
    samples: int = 0
    antenna_height: Optional[float] = None
    tilt: Optional[float] = None
    updated: str = ""
    operator: str = ""

    def to_dict(self) -> Dict[str, Any]:
        d = asdict(self)
        if not d["operator"] and d["mnc"]:
            d["operator"] = MNC_OPERATORS.get(d["mnc"], "")
        return d


# ---------------------------------------------------------------------------
# Database management
# ---------------------------------------------------------------------------

_SCHEMA = """
CREATE TABLE IF NOT EXISTS bts_stations (
    mcc         INTEGER NOT NULL DEFAULT 260,
    mnc         INTEGER NOT NULL,
    lac         INTEGER NOT NULL,
    cid         INTEGER NOT NULL,
    lat         REAL NOT NULL,
    lon         REAL NOT NULL,
    azimuth     REAL,
    range_m     INTEGER,
    radio       TEXT DEFAULT '',
    city        TEXT DEFAULT '',
    street      TEXT DEFAULT '',
    address     TEXT DEFAULT '',
    source      TEXT DEFAULT '',
    samples     INTEGER DEFAULT 0,
    antenna_height REAL,
    tilt        REAL,
    updated     TEXT DEFAULT '',
    PRIMARY KEY (mcc, mnc, lac, cid, source)
);

CREATE INDEX IF NOT EXISTS idx_bts_lac_cid ON bts_stations(lac, cid);
CREATE INDEX IF NOT EXISTS idx_bts_mnc_lac_cid ON bts_stations(mnc, lac, cid);
CREATE INDEX IF NOT EXISTS idx_bts_coords ON bts_stations(lat, lon);
CREATE INDEX IF NOT EXISTS idx_bts_city ON bts_stations(city);

CREATE TABLE IF NOT EXISTS bts_metadata (
    key   TEXT PRIMARY KEY,
    value TEXT
);
"""


class BTSDatabase:
    """SQLite-based BTS station database."""

    def __init__(self, db_path: Path):
        self.db_path = db_path
        self._conn: Optional[sqlite3.Connection] = None

    def _get_conn(self) -> sqlite3.Connection:
        if self._conn is None:
            self.db_path.parent.mkdir(parents=True, exist_ok=True)
            self._conn = sqlite3.connect(
                str(self.db_path), timeout=30, check_same_thread=False,
            )
            self._conn.row_factory = sqlite3.Row
            self._conn.execute("PRAGMA journal_mode=WAL")
            self._conn.execute("PRAGMA synchronous=NORMAL")
            self._conn.executescript(_SCHEMA)
        return self._conn

    def close(self):
        if self._conn:
            self._conn.close()
            self._conn = None

    # --- Lookup ---

    def lookup(
        self,
        lac: int,
        cid: int,
        mnc: Optional[int] = None,
        mcc: int = MCC_POLAND,
    ) -> List[BTSStation]:
        """Look up BTS stations by LAC + CID.

        Returns all matching stations (may have multiple sources or azimuths).
        """
        conn = self._get_conn()
        if mnc is not None:
            rows = conn.execute(
                "SELECT * FROM bts_stations WHERE mcc=? AND mnc=? AND lac=? AND cid=?",
                (mcc, mnc, lac, cid),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM bts_stations WHERE mcc=? AND lac=? AND cid=?",
                (mcc, lac, cid),
            ).fetchall()
        return [self._row_to_station(r) for r in rows]

    def lookup_best(
        self,
        lac: int,
        cid: int,
        mnc: Optional[int] = None,
        mcc: int = MCC_POLAND,
    ) -> Optional[BTSStation]:
        """Return single best station for LAC+CID.

        Priority: UKE (has azimuth) > billing > opencellid.
        """
        stations = self.lookup(lac, cid, mnc, mcc)
        if not stations:
            return None
        # Prefer UKE, then billing, then opencellid
        source_prio = {"uke": 0, "billing": 1, "opencellid": 2, "manual": 3}
        stations.sort(key=lambda s: source_prio.get(s.source, 99))
        return stations[0]

    def search_nearby(
        self,
        lat: float,
        lon: float,
        radius_deg: float = 0.01,
    ) -> List[BTSStation]:
        """Find stations near given coordinates (simple bounding box)."""
        conn = self._get_conn()
        rows = conn.execute(
            """SELECT * FROM bts_stations
               WHERE lat BETWEEN ? AND ? AND lon BETWEEN ? AND ?
               ORDER BY ABS(lat - ?) + ABS(lon - ?)
               LIMIT 50""",
            (lat - radius_deg, lat + radius_deg,
             lon - radius_deg, lon + radius_deg,
             lat, lon),
        ).fetchall()
        return [self._row_to_station(r) for r in rows]

    def get_stats(self) -> Dict[str, Any]:
        """Get database statistics."""
        conn = self._get_conn()
        total = conn.execute("SELECT COUNT(*) FROM bts_stations").fetchone()[0]
        by_source = {}
        for row in conn.execute(
            "SELECT source, COUNT(*) as cnt FROM bts_stations GROUP BY source"
        ).fetchall():
            by_source[row["source"]] = row["cnt"]
        by_radio = {}
        for row in conn.execute(
            "SELECT radio, COUNT(*) as cnt FROM bts_stations WHERE radio != '' GROUP BY radio"
        ).fetchall():
            by_radio[row["radio"]] = row["cnt"]
        cities = conn.execute(
            "SELECT COUNT(DISTINCT city) FROM bts_stations WHERE city != ''"
        ).fetchone()[0]
        return {
            "total_stations": total,
            "by_source": by_source,
            "by_radio": by_radio,
            "unique_cities": cities,
            "db_size_mb": round(self.db_path.stat().st_size / 1048576, 1)
            if self.db_path.exists() else 0,
        }

    # --- Import ---

    def import_opencellid_csv(
        self,
        csv_path: Path,
        mcc_filter: int = MCC_POLAND,
        progress_cb=None,
    ) -> int:
        """Import OpenCelliD CSV dump.

        CSV columns: radio, mcc, net, area, cell, unit, lon, lat,
                     range, samples, changeable, created, updated, averageSignal

        Args:
            csv_path: Path to the CSV file.
            mcc_filter: Only import stations with this MCC (260 for Poland).
            progress_cb: Optional callback(imported, total_lines).

        Returns:
            Number of imported stations.
        """
        conn = self._get_conn()
        imported = 0
        batch: List[tuple] = []
        batch_size = 5000
        total_lines = 0

        # Count lines for progress
        if progress_cb:
            with open(csv_path, "r", encoding="utf-8", errors="replace") as f:
                total_lines = sum(1 for _ in f) - 1  # minus header

        with open(csv_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if not header:
                return 0

            for i, row in enumerate(reader):
                if len(row) < 8:
                    continue
                try:
                    mcc = int(row[1])
                    if mcc != mcc_filter:
                        continue
                    mnc = int(row[2])
                    lac = int(row[3])
                    cid = int(row[4])
                    lon = float(row[6])
                    lat = float(row[7])
                    range_m = int(row[8]) if row[8] else None
                    samples = int(row[9]) if row[9] else 0
                    radio = row[0].upper()
                    updated = row[12] if len(row) > 12 else ""

                    batch.append((
                        mcc, mnc, lac, cid, lat, lon,
                        None, range_m, radio,
                        "", "", "", "opencellid",
                        samples, None, None, updated,
                    ))
                    imported += 1

                    if len(batch) >= batch_size:
                        self._insert_batch(conn, batch)
                        batch.clear()
                        if progress_cb:
                            progress_cb(imported, total_lines)

                except (ValueError, IndexError):
                    continue

        if batch:
            self._insert_batch(conn, batch)

        conn.commit()
        self._set_meta("opencellid_imported", str(int(time.time())))
        self._set_meta("opencellid_count", str(imported))
        log.info("Imported %d OpenCelliD stations (MCC=%d)", imported, mcc_filter)
        return imported

    def import_uke_csv(
        self,
        csv_path: Path,
        progress_cb=None,
    ) -> int:
        """Import UKE radio license CSV (legacy fallback).

        Delegates to the generic _import_uke_rows() after parsing CSV.
        """
        # Try to detect encoding and delimiter
        with open(csv_path, "rb") as f:
            sample = f.read(8192)

        encoding = "utf-8"
        if b"\xff\xfe" in sample[:4] or b"\xfe\xff" in sample[:4]:
            encoding = "utf-16"
        elif b"\xef\xbb\xbf" in sample[:4]:
            encoding = "utf-8-sig"
        else:
            try:
                sample.decode("utf-8")
            except UnicodeDecodeError:
                encoding = "cp1250"  # Polish Windows encoding

        delimiter = ";"
        if sample.count(b",") > sample.count(b";"):
            delimiter = ","

        with open(csv_path, "r", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            header = next(reader, None)
            if not header:
                return 0
            rows_list = [row for row in reader]

        return self._import_uke_rows(header, rows_list, progress_cb)

    def import_uke_xlsx(
        self,
        xlsx_path: Path,
        progress_cb=None,
    ) -> int:
        """Import UKE radio license XLSX file.

        UKE XLSX has columns including:
        - Długość geograficzna / StGeoDlg (longitude)
        - Szerokość geograficzna / StGeoSzr (latitude)
        - Azymut / IdAzymut (azimuth in degrees)
        - Kąt pochylenia / Tilt
        - Wysokość zawieszenia anteny / WysokoscAnteny
        - Operator / NazwaPodmiotu
        - Technology / Standard

        Returns number of imported stations.
        """
        try:
            import openpyxl
        except ImportError:
            log.warning("openpyxl not installed, trying csv fallback")
            return 0

        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
        total_imported = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)

            header_row = next(rows_iter, None)
            if not header_row:
                continue

            header = [str(c) if c is not None else "" for c in header_row]
            rows_list = []
            for row in rows_iter:
                rows_list.append([str(c) if c is not None else "" for c in row])

            if rows_list:
                count = self._import_uke_rows(header, rows_list, progress_cb, sheet_hint=sheet_name)
                total_imported += count

        wb.close()
        return total_imported

    def import_uke_zip(
        self,
        zip_path: Path,
        progress_cb=None,
    ) -> int:
        """Import UKE data from a ZIP archive containing XLSX files.

        UKE publishes BTS data as a ZIP with multiple XLSX files
        (one per technology/band: gsm900, lte800, umts2100, etc.)

        Returns total number of imported stations across all files.
        """
        import zipfile
        import tempfile

        if not zipfile.is_zipfile(str(zip_path)):
            log.warning("Not a valid ZIP file: %s", zip_path)
            return 0

        total_imported = 0
        tmp_dir = Path(tempfile.mkdtemp(prefix="uke_zip_"))

        try:
            with zipfile.ZipFile(str(zip_path), "r") as zf:
                # List all XLSX/CSV files in the archive
                members = [m for m in zf.namelist()
                          if not m.startswith("__MACOSX") and not m.startswith(".")
                          ]
                xlsx_files = [m for m in members if m.lower().endswith(".xlsx")]
                csv_files = [m for m in members if m.lower().endswith(".csv")]

                log.info("UKE ZIP contains %d XLSX and %d CSV files: %s",
                         len(xlsx_files), len(csv_files), members)

                # Extract all
                zf.extractall(str(tmp_dir))

            # Import XLSX files first
            for xlsx_name in xlsx_files:
                extracted = tmp_dir / xlsx_name
                if not extracted.exists():
                    continue
                try:
                    count = self.import_uke_xlsx(extracted, progress_cb)
                    log.info("UKE ZIP: %s → %d stations", xlsx_name, count)
                    total_imported += count
                except Exception as e:
                    log.warning("UKE ZIP: error importing %s: %s", xlsx_name, e)

            # Fallback: import CSV files if no XLSX found
            if not xlsx_files and csv_files:
                for csv_name in csv_files:
                    extracted = tmp_dir / csv_name
                    if not extracted.exists():
                        continue
                    try:
                        count = self.import_uke_csv(extracted, progress_cb)
                        log.info("UKE ZIP: %s → %d stations", csv_name, count)
                        total_imported += count
                    except Exception as e:
                        log.warning("UKE ZIP: error importing %s: %s", csv_name, e)

        finally:
            import shutil
            shutil.rmtree(tmp_dir, ignore_errors=True)

        conn = self._get_conn()
        conn.commit()
        self._set_meta("uke_imported", str(int(time.time())))
        self._set_meta("uke_count", str(total_imported))
        log.info("Imported %d total UKE stations from ZIP", total_imported)
        return total_imported

    def import_uke_file(
        self,
        file_path: Path,
        progress_cb=None,
    ) -> int:
        """Auto-detect UKE file format and import.

        Supports: .zip (with XLSX/CSV inside), .xlsx, .csv
        """
        suffix = file_path.suffix.lower()
        if suffix == ".zip":
            return self.import_uke_zip(file_path, progress_cb)
        elif suffix == ".xlsx":
            return self.import_uke_xlsx(file_path, progress_cb)
        elif suffix in (".csv", ".txt"):
            return self.import_uke_csv(file_path, progress_cb)
        else:
            log.warning("UKE import: unsupported file format: %s", suffix)
            return 0

    def _import_uke_rows(
        self,
        header: List[str],
        rows: List[List[str]],
        progress_cb=None,
        sheet_hint: str = "",
    ) -> int:
        """Common UKE import logic for both CSV and XLSX sources.

        Args:
            header: Column header names.
            rows: List of row data (each row is a list of strings).
            progress_cb: Optional progress callback(imported, total).
            sheet_hint: Sheet/file name for radio tech auto-detection.

        Returns number of imported stations.
        """
        conn = self._get_conn()
        imported = 0
        batch: List[tuple] = []
        batch_size = 5000

        col_map = self._map_uke_columns(header)
        if "lat" not in col_map or "lon" not in col_map:
            log.warning("UKE: could not find lat/lon columns in header: %s", header[:20])
            return 0

        # Auto-detect radio technology from sheet/file name
        radio_from_name = ""
        hint = sheet_hint.lower()
        if "lte" in hint or "4g" in hint:
            radio_from_name = "LTE"
        elif "umts" in hint or "3g" in hint:
            radio_from_name = "UMTS"
        elif "gsm" in hint or "2g" in hint:
            radio_from_name = "GSM"
        elif "nr" in hint or "5g" in hint:
            radio_from_name = "NR"
        elif "cdma" in hint:
            radio_from_name = "CDMA"

        for i, row in enumerate(rows):
            try:
                lat = self._parse_coord(
                    row[col_map["lat"]] if col_map.get("lat") is not None and col_map["lat"] < len(row) else ""
                )
                lon = self._parse_coord(
                    row[col_map["lon"]] if col_map.get("lon") is not None and col_map["lon"] < len(row) else ""
                )
                if not lat or not lon:
                    continue

                # Validate coords are in Poland range
                if not (48.0 < lat < 55.5 and 13.5 < lon < 25.0):
                    continue

                azimuth = None
                if col_map.get("azimuth") is not None and col_map["azimuth"] < len(row):
                    az_str = row[col_map["azimuth"]].strip()
                    if az_str and az_str not in ("-", "", "None"):
                        try:
                            azimuth = float(az_str.replace(",", "."))
                        except ValueError:
                            pass

                tilt = None
                if col_map.get("tilt") is not None and col_map["tilt"] < len(row):
                    tilt_str = row[col_map["tilt"]].strip()
                    if tilt_str and tilt_str not in ("-", "", "None"):
                        try:
                            tilt = float(tilt_str.replace(",", "."))
                        except ValueError:
                            pass

                antenna_height = None
                if col_map.get("height") is not None and col_map["height"] < len(row):
                    h_str = row[col_map["height"]].strip()
                    if h_str and h_str not in ("-", "", "None"):
                        try:
                            antenna_height = float(h_str.replace(",", "."))
                        except ValueError:
                            pass

                # Detect MNC from operator column
                mnc = 0
                if col_map.get("operator") is not None and col_map["operator"] < len(row):
                    op_text = row[col_map["operator"]].lower()
                    if "orange" in op_text or "ptk" in op_text:
                        mnc = 3
                    elif "t-mobile" in op_text or "ptc" in op_text:
                        mnc = 2
                    elif "play" in op_text or "p4" in op_text:
                        mnc = 6
                    elif "plus" in op_text or "polkomtel" in op_text:
                        mnc = 1

                # UKE doesn't have CID/LAC — use synthetic ID
                synthetic_cid = i + 1
                synthetic_lac = 0

                # Radio tech from column or sheet name
                radio = radio_from_name
                if col_map.get("radio") is not None and col_map["radio"] < len(row):
                    r_text = row[col_map["radio"]].upper()
                    if "LTE" in r_text or "4G" in r_text:
                        radio = "LTE"
                    elif "UMTS" in r_text or "3G" in r_text:
                        radio = "UMTS"
                    elif "GSM" in r_text or "2G" in r_text:
                        radio = "GSM"
                    elif "NR" in r_text or "5G" in r_text:
                        radio = "NR"

                # City from column if available
                city = ""
                if col_map.get("city") is not None and col_map["city"] < len(row):
                    city = row[col_map["city"]].strip()

                batch.append((
                    MCC_POLAND, mnc, synthetic_lac, synthetic_cid,
                    lat, lon, azimuth, None, radio,
                    city, "", "", "uke",
                    0, antenna_height, tilt, "",
                ))
                imported += 1

                if len(batch) >= batch_size:
                    self._insert_batch(conn, batch)
                    batch.clear()
                    if progress_cb:
                        progress_cb(imported, len(rows))

            except (ValueError, IndexError):
                continue

        if batch:
            self._insert_batch(conn, batch)

        conn.commit()
        if not sheet_hint:
            # Only set metadata from top-level call
            self._set_meta("uke_imported", str(int(time.time())))
            self._set_meta("uke_count", str(imported))
        log.info("Imported %d UKE stations%s", imported,
                 f" from {sheet_hint}" if sheet_hint else "")
        return imported

    def import_from_billing_records(
        self,
        records: List[Dict[str, Any]],
    ) -> int:
        """Import BTS data extracted from billing records (e.g. T-Mobile).

        T-Mobile billing records contain BTS X, BTS Y, CI, LAC, Azimuth, city, street.
        """
        conn = self._get_conn()
        imported = 0
        batch: List[tuple] = []

        for r in records:
            extra = r.get("extra", {})
            bts_x = extra.get("bts_x", "")
            bts_y = extra.get("bts_y", "")
            ci = r.get("location_cell_id", "")
            lac = r.get("location_lac", "")

            if not ci or not lac:
                continue
            if not bts_x and not bts_y:
                continue

            try:
                lon = float(bts_x) if bts_x else 0.0
                lat = float(bts_y) if bts_y else 0.0
                cid_int = int(ci)
                lac_int = int(lac)
            except (ValueError, TypeError):
                continue

            if lat == 0.0 and lon == 0.0:
                continue

            azimuth = None
            az_str = extra.get("azimuth", "")
            if az_str:
                try:
                    azimuth = float(az_str)
                except (ValueError, TypeError):
                    pass

            city = extra.get("bts_city", "") if "bts_city" in extra else ""
            street = extra.get("bts_street", "") if "bts_street" in extra else ""

            # Determine MNC from network operator
            mnc = 2  # T-Mobile default
            location = r.get("location", "")

            batch.append((
                MCC_POLAND, mnc, lac_int, cid_int,
                lat, lon, azimuth, None, "",
                city, street, f"{city}, {street}" if city and street else city or street,
                "billing", 1, None, None, "",
            ))
            imported += 1

        if batch:
            self._insert_batch(conn, batch)
            conn.commit()

        log.info("Imported %d BTS stations from billing records", imported)
        return imported

    def add_station(self, station: BTSStation) -> None:
        """Add or update a single station."""
        conn = self._get_conn()
        self._insert_batch(conn, [(
            station.mcc, station.mnc, station.lac, station.cid,
            station.lat, station.lon, station.azimuth, station.range_m,
            station.radio, station.city, station.street, station.address,
            station.source, station.samples, station.antenna_height,
            station.tilt, station.updated,
        )])
        conn.commit()

    def clear(self, source: Optional[str] = None) -> int:
        """Clear all stations or only from a specific source."""
        conn = self._get_conn()
        if source:
            cursor = conn.execute(
                "DELETE FROM bts_stations WHERE source=?", (source,)
            )
        else:
            cursor = conn.execute("DELETE FROM bts_stations")
        conn.commit()
        return cursor.rowcount

    # --- Internal helpers ---

    def _insert_batch(self, conn: sqlite3.Connection, batch: List[tuple]) -> None:
        conn.executemany(
            """INSERT OR REPLACE INTO bts_stations
               (mcc, mnc, lac, cid, lat, lon, azimuth, range_m, radio,
                city, street, address, source, samples, antenna_height, tilt, updated)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            batch,
        )

    def _set_meta(self, key: str, value: str) -> None:
        conn = self._get_conn()
        conn.execute(
            "INSERT OR REPLACE INTO bts_metadata (key, value) VALUES (?, ?)",
            (key, value),
        )
        conn.commit()

    def _get_meta(self, key: str) -> Optional[str]:
        conn = self._get_conn()
        row = conn.execute(
            "SELECT value FROM bts_metadata WHERE key=?", (key,)
        ).fetchone()
        return row[0] if row else None

    @staticmethod
    def _row_to_station(row: sqlite3.Row) -> BTSStation:
        return BTSStation(
            mcc=row["mcc"],
            mnc=row["mnc"],
            lac=row["lac"],
            cid=row["cid"],
            lat=row["lat"],
            lon=row["lon"],
            azimuth=row["azimuth"],
            range_m=row["range_m"],
            radio=row["radio"] or "",
            city=row["city"] or "",
            street=row["street"] or "",
            address=row["address"] or "",
            source=row["source"] or "",
            samples=row["samples"] or 0,
            antenna_height=row["antenna_height"],
            tilt=row["tilt"],
            updated=row["updated"] or "",
        )

    @staticmethod
    def _map_uke_columns(header: List[str]) -> Dict[str, int]:
        """Map UKE CSV/XLSX column names to indices.

        Handles various UKE naming conventions across years:
        - Polish long names: 'Długość geograficzna', 'Szerokość geograficzna'
        - UKE short codes: 'StGeoDlg', 'StGeoSzr', 'IdAzymut'
        - Various casing and spelling variants
        """
        col_map: Dict[str, int] = {}
        for i, h in enumerate(header):
            h_lower = h.strip().lower()
            h_stripped = h.strip()

            # Longitude
            if (re.search(r"d[łl]ugo[śs][ćc]\s*geogr", h_lower)
                    or h_lower in ("lon", "longitude", "dlugosc", "dlug_geo")
                    or h_stripped.lower() in ("stgeodlg", "dlugosc_geograficzna")):
                col_map["lon"] = i
            # Latitude
            elif (re.search(r"szeroko[śs][ćc]\s*geogr", h_lower)
                  or h_lower in ("lat", "latitude", "szerokosc", "szer_geo")
                  or h_stripped.lower() in ("stgeoszr", "szerokosc_geograficzna")):
                col_map["lat"] = i
            # Azimuth
            elif (re.search(r"azymut", h_lower)
                  or h_stripped.lower() in ("idazymut", "azimuth")):
                col_map["azimuth"] = i
            # Tilt
            elif (re.search(r"k[aą]t\s*pochyl", h_lower)
                  or h_lower in ("tilt", "pochylenie")
                  or h_stripped.lower() == "katpochylenia"):
                col_map["tilt"] = i
            # Antenna height
            elif (re.search(r"wysoko[śs][ćc].*anten", h_lower)
                  or h_lower in ("wysokosc_anteny", "wys_anteny")
                  or h_stripped.lower() in ("wysokoscanteny", "wysokosc")):
                col_map["height"] = i
            # Operator name
            elif (re.search(r"operator|nazwa\s*podmiotu", h_lower)
                  or h_stripped.lower() in ("nazwapodmiotu", "operator", "podmiot")):
                col_map["operator"] = i
            # Radio technology
            elif (re.search(r"system|technolog|standard", h_lower)
                  or h_stripped.lower() in ("standard", "technologia")):
                col_map["radio"] = i
            # City / location name
            elif (re.search(r"miejscowo[śs][ćc]|miasto|lokalizacja", h_lower)
                  or h_stripped.lower() in ("miejscowosc", "miasto", "stmiejscowosc")):
                col_map["city"] = i
        return col_map

    @staticmethod
    def _parse_coord(text: str) -> Optional[float]:
        """Parse coordinate string (handles both decimal and DMS formats)."""
        text = text.strip().replace(",", ".")
        if not text or text == "-":
            return None
        try:
            return float(text)
        except ValueError:
            # Try DMS format: 52°13'24.5"N
            m = re.match(
                r"(\d+)[°]\s*(\d+)[′']\s*([\d.]+)[″\"]?\s*([NSEW])?",
                text,
            )
            if m:
                deg = float(m.group(1)) + float(m.group(2)) / 60 + float(m.group(3)) / 3600
                if m.group(4) in ("S", "W"):
                    deg = -deg
                return deg
            return None


# ---------------------------------------------------------------------------
# Module-level convenience
# ---------------------------------------------------------------------------

_default_db: Optional[BTSDatabase] = None


def get_bts_db(data_dir: Optional[Path] = None) -> BTSDatabase:
    """Get or create the default BTS database instance."""
    global _default_db
    if _default_db is not None:
        return _default_db

    if data_dir is None:
        import os
        data_dir = Path(os.environ.get("AISTATEWEB_DATA_DIR", "data_www"))

    db_path = data_dir / "gsm" / "bts.db"
    _default_db = BTSDatabase(db_path)
    return _default_db
