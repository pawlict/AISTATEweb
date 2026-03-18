"""GSM subscriber identification file parsers.

Parses identification (subscriber data) files from Polish operators:
- Orange:   XLSX with "7. DANE OSOBOWE" header
- Play:     CSV (semicolon-delimited, cp1250 encoding, ="" quoting)
- Plus:     CSV (comma-delimited, utf-8-sig encoding)
- T-Mobile: CSV/XLSX (semicolon-delimited, metadata header + data rows)

Each parser normalises phone numbers and extracts:
  number, first_name, last_name, name, pesel, nip, regon, address, city,
  operator info, type, activation/deactivation dates, sim, imsi, document number.

The IdentificationStore keeps ALL records per MSISDN (multi-record) to support:
  - Temporal history (operator changes, address changes)
  - Cross-operator merging (same number at different operators)
  - PESEL-based grouping (same person with multiple numbers)

Usage:
    store = IdentificationStore()
    store.load_file(path)                 # auto-detect and parse
    info = store.lookup("48501234567")    # best SubscriberIdentification or None
    all_records = store.lookup_all("48501234567")  # full history
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

log = logging.getLogger("aistate.gsm.identification")


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class SubscriberIdentification:
    """Identification data for a single phone number."""

    number: str = ""           # normalised MSISDN  e.g. "48501234567"
    name: str = ""             # imię i nazwisko / nazwa firmy (original)
    first_name: str = ""       # wyodrębnione imię
    last_name: str = ""        # wyodrębnione nazwisko
    pesel: str = ""
    nip: str = ""
    regon: str = ""
    address: str = ""          # full address string
    city: str = ""
    document_number: str = ""
    document_type: str = ""
    subscriber_type: str = ""  # I=indywidualny, B=biznesowy, post-paid, pre-paid etc.
    operator: str = ""         # operator name / code
    activation_date: str = ""
    deactivation_date: str = ""
    sim: str = ""              # SIM ICCID
    imsi: str = ""
    service_type: str = ""     # e.g. NUMERY HURTOWE, Telefonia mobilna
    tariff: str = ""
    status: str = ""           # Aktywna, Nie znaleziono, etc.
    notes: str = ""            # extra info / uwagi
    email: str = ""            # email kontaktowy
    contract_type: str = ""    # Postpaid / Prepaid / MVNO
    contract_status: str = ""  # N (normalny), inne
    correspondence_address: str = ""  # adres korespondencyjny
    source_file: str = ""      # which file this came from
    source_operator: str = ""  # orange / play / plus / tmobile

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    # Phrases indicating "no data" / "not found" / "different operator"
    _NO_DATA_PHRASES = (
        "nie znaleziono", "nie był przydzielony", "nie obsługiwany",
        "numer nie obsługiwany", "brak danych", "brak informacji",
        "nie zidentyfikowano", "numer nieaktywny",
        "obsługiwany przez", "przeniesiony",
    )

    @property
    def has_owner_data(self) -> bool:
        """True if we have actual owner identification (not just 'not found')."""
        if not self.name:
            return False
        name_lower = self.name.lower().strip()
        return not any(ph in name_lower for ph in self._NO_DATA_PHRASES)

    @property
    def identification_type(self) -> str:
        """Type of identification result: 'person', 'company', 'other_operator', 'not_found', 'unknown'."""
        if not self.name:
            return "unknown"
        name_lower = self.name.lower().strip()

        # Check if it's a "different operator" message
        if "obsługiwany przez" in name_lower or "przeniesiony" in name_lower:
            return "other_operator"

        # Check for "not found" / "not assigned"
        if any(ph in name_lower for ph in self._NO_DATA_PHRASES):
            return "not_found"

        # Person vs company: companies have legal form indicators
        company_indicators = (
            "sp. z o.o", "spółka", "s.a.", "sp.j.", "sp.k.",
            "fundacja", "stowarzyszenie", "urząd", "szkoła",
            "zakład", "przedsiębiorstwo", "firma", "gospodarstwo",
            "sp. zo.o", "bank ", "poczta polska", "centrum ",
        )
        if any(ind in name_lower for ind in company_indicators):
            return "company"

        # If has PESEL → person
        if self.pesel:
            return "person"

        # If has NIP/REGON → company
        if self.nip or self.regon:
            return "company"

        # Default: assume person if name looks like "Imie Nazwisko"
        words = self.name.strip().split()
        if 2 <= len(words) <= 4 and all(w[0].isupper() for w in words if w):
            return "person"

        return "unknown"

    @property
    def display_label(self) -> str:
        """Short display string for UI table column."""
        id_type = self.identification_type

        if id_type == "other_operator":
            # Extract operator name from message
            m = re.search(r'obsługiwany przez\s*[-–—]?\s*(.+)', self.name, re.IGNORECASE)
            if m:
                op_name = m.group(1).strip().rstrip(".")
                return f"[inny oper.: {op_name}]"
            if "przeniesiony" in self.name.lower():
                return "[przeniesiony do innego oper.]"
            return "[inny operator]"

        if id_type == "not_found":
            return "[brak danych]"

        parts = []
        if self.last_name and self.first_name:
            parts.append(f"{self.first_name} {self.last_name}")
        elif self.name:
            parts.append(self.name)
        if self.city:
            parts.append(self.city)
        return ", ".join(parts) if parts else self.status or "[brak danych]"


# ---------------------------------------------------------------------------
# Phone number normalisation
# ---------------------------------------------------------------------------

_RE_NONDIGIT = re.compile(r"\D")


def normalise_msisdn(raw: str) -> str:
    """Normalise an MSISDN to digits-only, 11-digit (48xxx) format.

    Handles: +48XXXXXXXXX, 48XXXXXXXXX, 9-digit, ='48...' formats.
    """
    if not raw:
        return ""
    s = str(raw).strip()
    # Strip ="" quoting from Play CSVs
    s = s.strip("=").strip('"').strip("'")
    s = _RE_NONDIGIT.sub("", s)
    # Remove leading +
    if s.startswith("0048"):
        s = s[2:]
    if len(s) == 9:
        s = "48" + s
    return s


# ---------------------------------------------------------------------------
# Name parsing helpers
# ---------------------------------------------------------------------------

def _parse_name_parts(full_name: str, source_operator: str = "") -> Tuple[str, str]:
    """Extract (first_name, last_name) from a full name string.

    Handles various operator formats:
    - T-Mobile:  "NAZWISKO,IMIE" or "NAZWISKO IMIE" or "IMIE NAZWISKO"
    - Orange:    "NAZWISKO,IMIE" (with comma) or "IMIE NAZWISKO"
    - Play:      separate IMIE/NAZWISKO columns — pass directly
    - Plus:      "IMIE NAZWISKO" or "NAZWA FIRMY"
    """
    if not full_name or not full_name.strip():
        return ("", "")

    name = full_name.strip()

    # Skip company names and status messages
    name_lower = name.lower()
    skip_indicators = (
        "sp. z o.o", "spółka", "s.a.", "sp.j.", "sp.k.",
        "fundacja", "stowarzyszenie", "urząd", "szkoła",
        "zakład", "przedsiębiorstwo", "firma", "gospodarstwo",
        "sp. zo.o", "bank ", "poczta polska", "centrum ",
        "nie znaleziono", "nie był", "przeniesiony", "brak danych",
        "obsługiwany", "biuro obsługi",
    )
    if any(ind in name_lower for ind in skip_indicators):
        return ("", "")

    # T-Mobile / Orange comma format: "NAZWISKO,IMIE" or "NAZWISKO,IMIE DRUGIE"
    if "," in name:
        parts = name.split(",", 1)
        last = parts[0].strip()
        first = parts[1].strip() if len(parts) > 1 else ""
        return (_title_case(first), _title_case(last))

    # Space-separated: need to figure out which is first/last
    words = name.split()
    if len(words) < 2:
        # Single word — ambiguous, treat as last name
        return ("", _title_case(name))

    if len(words) == 2:
        # Two words: format depends on operator
        # T-Mobile/Orange (without comma): "NAZWISKO IMIE" → first word = last name
        # Plus: "IMIE NAZWISKO" → first word = first name (header: "Imię i Nazwisko/Nazwa")
        if source_operator in ("plus",):
            return (_title_case(words[0]), _title_case(words[1]))
        # Default (T-Mobile, Orange, unknown): NAZWISKO IMIE
        return (_title_case(words[1]), _title_case(words[0]))

    if len(words) == 3:
        # Three words: "NAZWISKO IMIE DRUGIE" (T-Mobile, Orange)
        # or "IMIE DRUGIE NAZWISKO" (Plus)
        if source_operator in ("plus",):
            return (_title_case(" ".join(words[:-1])), _title_case(words[-1]))
        # Default: first word = last name
        return (_title_case(" ".join(words[1:])), _title_case(words[0]))

    if len(words) == 4:
        # Four words: e.g. "IMIE DRUGIE NAZWISKO1 NAZWISKO2"
        return (_title_case(" ".join(words[:2])), _title_case(" ".join(words[2:])))

    # 5+ words — likely company or complex name, skip parsing
    return ("", "")


def _title_case(s: str) -> str:
    """Convert UPPERCASE name to Title Case, handling Polish names."""
    if not s:
        return ""
    # If already mixed case, keep as-is
    if not s.isupper():
        return s.strip()
    # Title-case each word
    return " ".join(w.capitalize() for w in s.split())


# ---------------------------------------------------------------------------
# Orange parser (XLSX)
# ---------------------------------------------------------------------------

def _parse_orange_xlsx(file_path: Path) -> List[SubscriberIdentification]:
    """Parse Orange '7. DANE OSOBOWE' identification XLSX.

    Structure:
        Row 1:  "7. DANE OSOBOWE"
        Row 7:  Header: LP | IDENTYFIKATOR USŁUGI | OPER | RU | RU2 |
                DATA AKT. | DATA DEZAKT. | TYP | ABONENT | PESEL/REGON/NIP |
                KOD | MIASTO | ULICA NR | KOD_2 | MIASTO_2 | ULICA NR_2 | NUMER GŁÓWNY
        Row 8+: Data rows until "KONIEC WYKAZU"
    """
    import openpyxl

    results: List[SubscriberIdentification] = []
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row (contains "IDENTYFIKATOR USŁUGI")
        header_idx = None
        for i, row in enumerate(rows):
            cells = [str(c or "").strip().upper() for c in row]
            if any("IDENTYFIKATOR" in c for c in cells):
                header_idx = i
                break

        if header_idx is None:
            log.warning("Orange ID: header row not found in sheet %s", sheet_name)
            continue

        header = [str(c or "").strip().upper() for c in rows[header_idx]]

        # Map column indices
        col_map = {}
        for ci, h in enumerate(header):
            if "IDENTYFIKATOR" in h:
                col_map["msisdn"] = ci
            elif h == "OPER":
                col_map["oper"] = ci
            elif h == "RU":
                col_map["ru"] = ci
            elif h.startswith("DATA AKT"):
                col_map["activation"] = ci
            elif h.startswith("DATA DEZAKT"):
                col_map["deactivation"] = ci
            elif h == "TYP":
                col_map["type"] = ci
            elif h == "ABONENT":
                col_map["name"] = ci
            elif "PESEL" in h:
                col_map["pesel_nip"] = ci
            elif h == "KOD":
                col_map["zip"] = ci
            elif h == "MIASTO" and "miasto" not in col_map:
                col_map["miasto"] = ci
            elif h.startswith("ULICA"):
                if "street" not in col_map:
                    col_map["street"] = ci
            elif h == "NUMER G" or "GŁÓWNY" in h or "GLOWNY" in h:
                col_map["main_number"] = ci

        def _cell(row, key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        # Parse data rows
        for row in rows[header_idx + 1:]:
            cells_str = [str(c or "").strip() for c in row]
            # Stop at "KONIEC WYKAZU"
            if any("KONIEC" in c.upper() for c in cells_str[:3]):
                break
            # Skip empty rows
            msisdn_raw = _cell(row, "msisdn")
            if not msisdn_raw:
                continue

            msisdn = normalise_msisdn(msisdn_raw)
            if not msisdn:
                continue

            name = _cell(row, "name")
            pesel_nip = _cell(row, "pesel_nip")
            city = _cell(row, "miasto")
            street = _cell(row, "street")
            zipcode = _cell(row, "zip")

            # Build address
            addr_parts = []
            if street:
                addr_parts.append(street)
            if zipcode:
                addr_parts.append(zipcode)
            if city:
                addr_parts.append(city)
            address = ", ".join(addr_parts)

            # Parse PESEL/NIP/REGON
            pesel, nip, regon = "", "", ""
            if pesel_nip:
                pn = _RE_NONDIGIT.sub("", str(pesel_nip))
                if len(pn) == 11:
                    pesel = pn
                elif len(pn) == 10:
                    nip = pn
                elif len(pn) == 9 or len(pn) == 14:
                    regon = pn

            sub_type = _cell(row, "type")

            # Parse name parts
            first_name, last_name = _parse_name_parts(name, "orange")

            results.append(SubscriberIdentification(
                number=msisdn,
                name=name,
                first_name=first_name,
                last_name=last_name,
                pesel=pesel,
                nip=nip,
                regon=regon,
                address=address,
                city=city,
                subscriber_type=sub_type,
                operator=_cell(row, "oper"),
                activation_date=_cell(row, "activation"),
                deactivation_date=_cell(row, "deactivation"),
                source_file=file_path.name,
                source_operator="orange",
            ))

    wb.close()
    return results


# ---------------------------------------------------------------------------
# Play parser (CSV, semicolon, cp1250, ="" quoting)
# ---------------------------------------------------------------------------

def _parse_play_csv(file_path: Path) -> List[SubscriberIdentification]:
    """Parse Play identification CSV.

    - Semicolon-delimited
    - Encoding: cp1250 (Polish Windows)
    - Values wrapped in ="" quoting: ="value"
    - Header: DATA_OD;DATA_DO;MSISDN_PSTN;SIM_NUMER;IMSI;PESEL;REGON;NIP;
              IMIE;NAZWISKO;NAZWA;ADRES;PIERWSZE_LOGOWANIE;UWAGI;USLUGA;TARYFA;
              ID_USLUGI;ID_LACZA;NR_DOKUMENTU;TYP_DOKUMENTU;ADRES_MAC;
              SKRZYNKA_POCZTOWA;TELEFON_KONTAKTOWY;EMAIL_KONTAKTOWY;
              DATA_I_GODZINA_LOG;BTS_X_LOG;BTS_Y_LOG;AZYM_LOG;BEAM_WIDTH_LOG;RANGE_LOG
    """
    results: List[SubscriberIdentification] = []

    # Try encodings
    content = None
    for enc in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            content = file_path.read_text(encoding=enc)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if content is None:
        log.error("Play ID: cannot decode %s", file_path)
        return results

    # Strip ="" quoting
    def strip_eq(val: str) -> str:
        v = val.strip()
        if v.startswith('="') and v.endswith('"'):
            return v[2:-1]
        if v.startswith("='") and v.endswith("'"):
            return v[2:-1]
        return v

    reader = csv.reader(io.StringIO(content), delimiter=";")
    header = None

    for row_idx, row in enumerate(reader):
        if not row:
            continue

        # Find header row
        if header is None:
            cleaned = [strip_eq(c).upper().strip() for c in row]
            if any("MSISDN" in c for c in cleaned):
                header = cleaned
                continue
            else:
                continue

        # Parse data row
        cells = [strip_eq(c) for c in row]

        def _col(name):
            """Get column value by header name (partial match)."""
            for i, h in enumerate(header):
                if name in h:
                    return cells[i] if i < len(cells) else ""
            return ""

        msisdn_raw = _col("MSISDN")
        if not msisdn_raw:
            continue
        msisdn = normalise_msisdn(msisdn_raw)
        if not msisdn:
            continue

        imie = _col("IMIE")
        nazwisko = _col("NAZWISKO")
        nazwa = _col("NAZWA")
        name = ""
        first_name = ""
        last_name = ""
        if imie or nazwisko:
            name = f"{imie} {nazwisko}".strip()
            first_name = _title_case(imie)
            last_name = _title_case(nazwisko)
        elif nazwa:
            name = nazwa
            first_name, last_name = _parse_name_parts(nazwa, "play")

        pesel = _col("PESEL")
        nip = _col("NIP")
        regon = _col("REGON")

        address_raw = _col("ADRES")
        # Play uses "Adres glowny: ... | Adres korespondencyjny: ..."
        address = address_raw.split("|")[0].strip()
        if address.lower().startswith("adres glowny:"):
            address = address[len("adres glowny:"):].strip()
        elif address.lower().startswith("adres główny:"):
            address = address[len("adres główny:"):].strip()

        # Extract correspondence address
        correspondence_address = ""
        if "|" in address_raw:
            for part in address_raw.split("|"):
                part_s = part.strip()
                lp = part_s.lower()
                if lp.startswith("adres korespondencyjny:"):
                    correspondence_address = part_s[len("adres korespondencyjny:"):].strip()
                    break

        # Extract city from address (last part after postal code)
        city = ""
        addr_match = re.search(r'\d{2}-\d{3}\s+(.+?)(?:\||$)', address_raw)
        if addr_match:
            city = addr_match.group(1).strip()

        sim = _col("SIM")
        imsi = _col("IMSI")
        service = _col("USLUGA")
        tariff = _col("TARYFA")
        doc_nr = _col("NR_DOKUMENTU")
        doc_type = _col("TYP_DOKUMENTU")
        uwagi = _col("UWAGI")
        data_od = _col("DATA_OD")
        data_do = _col("DATA_DO")
        email = _col("EMAIL")

        # Determine contract type from TARYFA/USLUGA
        contract_type = ""
        if tariff:
            tl = tariff.lower()
            if "prepaid" in tl or "pre-paid" in tl:
                contract_type = "Prepaid"
            elif "postpaid" in tl or "post-paid" in tl:
                contract_type = "Postpaid"
            elif "mvno" in tl:
                contract_type = "MVNO"

        results.append(SubscriberIdentification(
            number=msisdn,
            name=name,
            first_name=first_name,
            last_name=last_name,
            pesel=pesel,
            nip=nip,
            regon=regon,
            address=address,
            city=city,
            sim=sim,
            imsi=imsi,
            service_type=service,
            tariff=tariff,
            document_number=doc_nr,
            document_type=doc_type,
            notes=uwagi,
            email=email,
            contract_type=contract_type,
            correspondence_address=correspondence_address,
            activation_date=data_od,
            deactivation_date=data_do,
            source_file=file_path.name,
            source_operator="play",
        ))

    return results


# ---------------------------------------------------------------------------
# Plus parser (CSV, comma, utf-8-sig)
# ---------------------------------------------------------------------------

def _parse_plus_csv(file_path: Path) -> List[SubscriberIdentification]:
    """Parse Plus (Polkomtel) identification CSV.

    - Comma-delimited
    - Encoding: utf-8-sig
    - Header: Parametr,MSISDN,PESEL,SIM,Numer dokumentu,Typ dokumentu,NIP,
              Typ MSISDN,Ważne od,Ważne do,Imię i Nazwisko/Nazwa,Adres,Status
    """
    results: List[SubscriberIdentification] = []

    content = None
    for enc in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            content = file_path.read_text(encoding=enc)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    if content is None:
        log.error("Plus ID: cannot decode %s", file_path)
        return results

    reader = csv.reader(io.StringIO(content), delimiter=",")
    header = None

    for row_idx, row in enumerate(reader):
        if not row:
            continue

        # Find header row
        if header is None:
            cleaned = [c.strip().upper() for c in row]
            if any("MSISDN" in c for c in cleaned):
                header = cleaned
                continue
            else:
                continue

        cells = [c.strip() for c in row]

        def _col(name):
            for i, h in enumerate(header):
                if name in h:
                    return cells[i] if i < len(cells) else ""
            return ""

        msisdn_raw = _col("MSISDN")
        if not msisdn_raw:
            # Try "Parametr" column
            msisdn_raw = _col("PARAMETR")
        if not msisdn_raw:
            continue
        msisdn = normalise_msisdn(msisdn_raw)
        if not msisdn:
            continue

        name = _col("NAZWISKO") or _col("NAZWA")
        if not name:
            # Try "Imię i Nazwisko" combined column
            for i, h in enumerate(header):
                if "NAZWISK" in h or "NAZWA" in h:
                    name = cells[i] if i < len(cells) else ""
                    break

        # Parse name parts
        first_name, last_name = _parse_name_parts(name, "plus")

        pesel = _col("PESEL")
        nip = _col("NIP")
        sim = _col("SIM")
        doc_nr = _col("NUMER DOKUMENTU") or _col("NR_DOKUMENTU")
        doc_type = _col("TYP DOKUMENTU") or _col("TYP_DOKUMENTU")
        msisdn_type = _col("TYP MSISDN") or _col("TYP")
        valid_from = _col("OD")
        valid_to = _col("DO")
        address = _col("ADRES")
        status = _col("STATUS")

        # Extract city from address (e.g. "UL. TRAKTOROWA 126 91-204 ŁÓDŹ POL")
        city = ""
        addr_match = re.search(r'\d{2}-\d{3}\s+(.+?)(?:\s+POL\s*$|\s*$)', address, re.IGNORECASE)
        if addr_match:
            city = addr_match.group(1).strip()
            # Remove trailing country code
            city = re.sub(r'\s+POL\s*$', '', city, flags=re.IGNORECASE).strip()

        # Determine contract type from msisdn_type
        contract_type = ""
        if msisdn_type:
            mt = msisdn_type.lower()
            if "pre" in mt:
                contract_type = "Prepaid"
            elif "post" in mt:
                contract_type = "Postpaid"

        results.append(SubscriberIdentification(
            number=msisdn,
            name=name,
            first_name=first_name,
            last_name=last_name,
            pesel=pesel,
            nip=nip,
            address=address,
            city=city,
            sim=sim,
            document_number=doc_nr,
            document_type=doc_type,
            subscriber_type=msisdn_type,
            contract_type=contract_type,
            activation_date=valid_from,
            deactivation_date=valid_to,
            status=status,
            source_file=file_path.name,
            source_operator="plus",
        ))

    return results


# ---------------------------------------------------------------------------
# T-Mobile parser (CSV + XLSX)
# ---------------------------------------------------------------------------

def _read_tmobile_content(file_path: Path) -> Optional[List[List[str]]]:
    """Read T-Mobile identification file (CSV or XLSX) into a list of rows.

    T-Mobile provides identification in both CSV and XLSX formats with
    identical structure:
    - Metadata header (6 rows): zlecenie, zapytanie, sygnatura, daty, lista MSISDN
    - Empty row
    - Column header row
    - Data rows

    CSV: semicolon-delimited, values quoted with double-quotes
    XLSX: single sheet named after the file
    """
    suffix = file_path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c or "").strip() for c in row])
            wb.close()
            return rows
        except Exception as e:
            log.error("T-Mobile XLSX read error: %s", e)
            return None

    if suffix == ".csv":
        content = None
        for enc in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
            try:
                content = file_path.read_text(encoding=enc)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        if content is None:
            log.error("T-Mobile ID: cannot decode %s", file_path)
            return None

        rows = []
        reader = csv.reader(io.StringIO(content), delimiter=";")
        for row in reader:
            rows.append([c.strip().strip('"') for c in row])
        return rows

    return None


def _parse_tmobile(file_path: Path) -> List[SubscriberIdentification]:
    """Parse T-Mobile identification file (CSV or XLSX).

    Header columns:
        Status kontraktu | Typ kontraktu | SIM | IMSI | MSISDN |
        Aktywacja MSISDN | Wyłączenie MSISDN | Abonent | PESEL |
        Nr dokumentu tożsamości | Abonent od | Abonent do |
        Miejscowosc | Kod | Ulica | Nr |
        Koresp. miasto | Koresp. kod | Koresp. ulica | Koresp. nr |
        Koresp. od | Koresp. do | EMAIL | EMAIL Od | EMAIL Do
    """
    results: List[SubscriberIdentification] = []

    rows = _read_tmobile_content(file_path)
    if not rows:
        return results

    # Find header row — contains "MSISDN" and "Abonent" and "PESEL"
    header_idx = None
    for i, row in enumerate(rows):
        row_upper = [c.upper() for c in row]
        if any("MSISDN" in c for c in row_upper) and any("PESEL" in c for c in row_upper):
            header_idx = i
            break

    if header_idx is None:
        log.warning("T-Mobile ID: header row not found in %s", file_path)
        return results

    header = [c.strip().upper() for c in rows[header_idx]]

    # Build column index map
    col_map: Dict[str, int] = {}
    for ci, h in enumerate(header):
        h_norm = h.strip()
        if h_norm == "MSISDN":
            col_map["msisdn"] = ci
        elif "STATUS KONTRAKTU" in h_norm or h_norm == "STATUS KONTRAKTU":
            col_map["contract_status"] = ci
        elif "TYP KONTRAKTU" in h_norm:
            col_map["contract_type"] = ci
        elif h_norm == "SIM":
            col_map["sim"] = ci
        elif h_norm == "IMSI":
            col_map["imsi"] = ci
        elif "AKTYWACJA" in h_norm:
            col_map["activation"] = ci
        elif "WYLACZENIE" in h_norm or "WYŁĄCZENIE" in h_norm:
            col_map["deactivation"] = ci
        elif h_norm == "ABONENT":
            col_map["name"] = ci
        elif h_norm == "PESEL":
            col_map["pesel"] = ci
        elif "NR DOKUMENTU" in h_norm or "DOKUMENTU TO" in h_norm:
            col_map["document"] = ci
        elif h_norm == "ABONENT OD":
            col_map["subscriber_from"] = ci
        elif h_norm == "ABONENT DO":
            col_map["subscriber_to"] = ci
        elif h_norm == "MIEJSCOWOSC" or h_norm == "MIEJSCOWOŚĆ":
            if "city" not in col_map:
                col_map["city"] = ci
        elif h_norm == "KOD":
            if "zip" not in col_map:
                col_map["zip"] = ci
        elif h_norm == "ULICA":
            if "street" not in col_map:
                col_map["street"] = ci
        elif h_norm == "NR":
            if "street_nr" not in col_map:
                col_map["street_nr"] = ci
        elif "KORESP" in h_norm and "MIASTO" in h_norm:
            col_map["corr_city"] = ci
        elif "KORESP" in h_norm and "KOD" in h_norm:
            col_map["corr_zip"] = ci
        elif "KORESP" in h_norm and "ULICA" in h_norm:
            col_map["corr_street"] = ci
        elif "KORESP" in h_norm and "NR" in h_norm and "ULICA" not in h_norm:
            col_map["corr_nr"] = ci
        elif h_norm == "EMAIL":
            col_map["email"] = ci

    def _cell(row_data, key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row_data):
            return ""
        return str(row_data[idx] or "").strip()

    # Parse data rows
    for row in rows[header_idx + 1:]:
        if not row or all(not c.strip() for c in row):
            continue

        msisdn_raw = _cell(row, "msisdn")
        if not msisdn_raw:
            continue
        msisdn = normalise_msisdn(msisdn_raw)
        if not msisdn:
            continue

        name = _cell(row, "name")
        pesel = _cell(row, "pesel")
        # PESEL may be numeric in XLSX — ensure string
        pesel = _RE_NONDIGIT.sub("", str(pesel)) if pesel else ""
        if pesel and len(pesel) != 11:
            pesel = ""

        document = _cell(row, "document")
        city = _cell(row, "city")
        zipcode = _cell(row, "zip")
        street = _cell(row, "street")
        street_nr = _cell(row, "street_nr")

        # Build address
        addr_parts = []
        if street:
            s = street
            if street_nr:
                s += " " + street_nr
            addr_parts.append(s)
        if zipcode:
            addr_parts.append(zipcode)
        if city:
            addr_parts.append(city)
        address = ", ".join(addr_parts)

        # Build correspondence address
        corr_city = _cell(row, "corr_city")
        corr_zip = _cell(row, "corr_zip")
        corr_street = _cell(row, "corr_street")
        corr_nr = _cell(row, "corr_nr")
        corr_parts = []
        if corr_street:
            cs = corr_street
            if corr_nr:
                cs += " " + corr_nr
            corr_parts.append(cs)
        if corr_zip:
            corr_parts.append(corr_zip)
        if corr_city:
            corr_parts.append(corr_city)
        correspondence_address = ", ".join(corr_parts)

        contract_status = _cell(row, "contract_status")
        contract_type = _cell(row, "contract_type")
        sim = _cell(row, "sim")
        imsi = _cell(row, "imsi")
        activation = _cell(row, "activation")
        deactivation = _cell(row, "deactivation")
        email = _cell(row, "email")

        # Parse name parts — T-Mobile uses "NAZWISKO,IMIE" or "NAZWISKO IMIE"
        first_name, last_name = _parse_name_parts(name, "tmobile")

        # Determine subscriber_type from contract_status TYP field
        # T-Mobile has "I" (indywidualny) and "B" (biznesowy) in original data
        # but in our test data the TYP is in a separate position
        sub_type = ""
        if contract_type:
            sub_type = contract_type

        results.append(SubscriberIdentification(
            number=msisdn,
            name=name,
            first_name=first_name,
            last_name=last_name,
            pesel=pesel,
            document_number=document,
            address=address,
            city=city,
            correspondence_address=correspondence_address,
            subscriber_type=sub_type,
            contract_type=contract_type,
            contract_status=contract_status,
            sim=sim,
            imsi=imsi,
            email=email,
            activation_date=activation,
            deactivation_date=deactivation,
            source_file=file_path.name,
            source_operator="tmobile",
        ))

    return results


# ---------------------------------------------------------------------------
# Auto-detection and store
# ---------------------------------------------------------------------------

def detect_id_format(file_path: Path) -> Optional[str]:
    """Detect the operator format of an identification file.

    Returns: "orange", "play", "plus", "tmobile", or None.
    """
    suffix = file_path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                found_texts = []
                for row in ws.iter_rows(max_row=10, values_only=True):
                    text = " ".join(str(c or "") for c in row).upper()
                    found_texts.append(text)
                combined = " ".join(found_texts)

                # Orange: has "DANE OSOBOWE" or "IDENTYFIKATOR USŁUGI"
                if "DANE OSOBOWE" in combined or "IDENTYFIKATOR USŁUGI" in combined:
                    wb.close()
                    return "orange"

                # T-Mobile: has "Abonent dla MSISDN" or "PESEL" + "Abonent" header
                if "ABONENT DLA MSISDN" in combined:
                    wb.close()
                    return "tmobile"
                if "STATUS KONTRAKTU" in combined and "PESEL" in combined:
                    wb.close()
                    return "tmobile"
            wb.close()
        except Exception as e:
            log.debug("detect_id_format XLSX error: %s", e)
        return None

    if suffix == ".csv":
        # Read file content to detect format
        content = None
        for enc in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
            try:
                content = file_path.read_text(encoding=enc, errors="replace")
                break
            except Exception:
                continue
        if not content:
            return None

        # Use a larger window for T-Mobile (metadata MSISDN list can be very long)
        first_lines = content[:3000].upper()
        # Also check deeper content for header detection
        extended_lines = content[:20000].upper()

        # T-Mobile: semicolon-delimited, metadata has "Abonent dla MSISDN" or
        # "Identyfikator zlecenia" + later "PESEL" column header
        # Note: T-Mobile metadata MSISDN list can be >3000 chars,
        # so check metadata markers in first_lines but PESEL in extended
        if "ABONENT DLA MSISDN" in first_lines:
            if "PESEL" in extended_lines:
                return "tmobile"

        if ("IDENTYFIKATOR ZLECENIA" in first_lines
                and ";" in first_lines[:500]):
            # Check for PESEL/Abonent header deeper in the file
            if "PESEL" in extended_lines and "ABONENT" in extended_lines:
                return "tmobile"

        if "STATUS KONTRAKTU" in extended_lines and "PESEL" in extended_lines:
            return "tmobile"

        # Play: semicolon-delimited, has MSISDN_PSTN
        if "MSISDN_PSTN" in first_lines and ";" in first_lines[:500]:
            return "play"

        # Plus: comma-delimited, has "Typ MSISDN" or "Ważne od"
        if ("TYP MSISDN" in first_lines or "WAŻNE OD" in first_lines
                or "WAZNE OD" in first_lines):
            return "plus"

        # Fallback: check for MSISDN header with comma delimiter
        # (but not if it looks like T-Mobile metadata with semicolons)
        if "MSISDN" in first_lines and "," in first_lines[:500]:
            # Avoid misdetecting T-Mobile CSV as Plus
            if ";" not in first_lines[:200]:
                return "plus"
            # If semicolons come first, it's not Plus
            first_semi = first_lines.find(";")
            first_comma = first_lines.find(",")
            if first_comma < first_semi:
                return "plus"

        # Fallback: check for semicolon-delimited with MSISDN
        if "MSISDN" in first_lines and ";" in first_lines[:500]:
            return "play"

    return None


def parse_identification_file(file_path: Path) -> List[SubscriberIdentification]:
    """Auto-detect format and parse an identification file.

    Returns list of SubscriberIdentification records.
    """
    fmt = detect_id_format(file_path)
    if fmt is None:
        log.warning("Cannot detect format of identification file: %s", file_path)
        return []

    log.info("Parsing identification file: %s (format: %s)", file_path.name, fmt)

    if fmt == "orange":
        return _parse_orange_xlsx(file_path)
    elif fmt == "play":
        return _parse_play_csv(file_path)
    elif fmt == "plus":
        return _parse_plus_csv(file_path)
    elif fmt == "tmobile":
        return _parse_tmobile(file_path)
    else:
        return []


# ---------------------------------------------------------------------------
# Record scoring for smart selection
# ---------------------------------------------------------------------------

def _record_score(rec: SubscriberIdentification) -> int:
    """Score a record for quality — higher is better.

    Used to pick the best record when multiple exist for an MSISDN.
    """
    score = 0
    if rec.has_owner_data:
        score += 100
    if rec.pesel:
        score += 50
    if rec.first_name and rec.last_name:
        score += 30
    elif rec.name:
        score += 10
    if rec.city:
        score += 10
    if rec.address:
        score += 10
    if rec.document_number:
        score += 5
    if rec.nip:
        score += 5
    if rec.email:
        score += 3
    # Prefer active records
    if rec.deactivation_date:
        deact = rec.deactivation_date.strip()
        if deact.startswith("9999") or not deact:
            score += 20  # still active
    else:
        score += 20  # no deactivation = assumed active
    # Penalise "not found" / "other operator" records
    id_type = rec.identification_type
    if id_type == "not_found":
        score -= 200
    elif id_type == "other_operator":
        score -= 100
    return score


class IdentificationStore:
    """In-memory store of identification data for phone number lookups.

    Multi-record: keeps ALL records per MSISDN to support:
    - Temporal history (operator changes, address changes)
    - Cross-operator merging
    - PESEL-based grouping

    Usage:
        store = IdentificationStore()
        store.load_file(Path("orange_id.xlsx"))
        store.load_file(Path("play_id.csv"))
        info = store.lookup("48501234567")       # best record
        all_info = store.lookup_all("48501234567") # full history
    """

    def __init__(self):
        self._records: Dict[str, List[SubscriberIdentification]] = {}
        self._files_loaded: List[str] = []

    @property
    def count(self) -> int:
        """Total unique MSISDNs with at least one record."""
        return len(self._records)

    @property
    def total_records(self) -> int:
        """Total number of individual records across all MSISDNs."""
        return sum(len(recs) for recs in self._records.values())

    @property
    def files_loaded(self) -> List[str]:
        return list(self._files_loaded)

    def load_file(self, file_path: Path) -> int:
        """Parse and load an identification file. Returns number of records loaded."""
        records = parse_identification_file(file_path)
        count = 0
        for rec in records:
            if rec.number:
                if rec.number not in self._records:
                    self._records[rec.number] = []
                self._records[rec.number].append(rec)
                count += 1
        self._files_loaded.append(file_path.name)
        log.info("Loaded %d identification records from %s (total MSISDNs: %d)",
                 count, file_path.name, len(self._records))
        return count

    def load_files(self, file_paths: List[Path]) -> int:
        """Load multiple identification files."""
        total = 0
        for fp in file_paths:
            total += self.load_file(fp)
        return total

    def lookup(self, number: str) -> Optional[SubscriberIdentification]:
        """Look up best identification by phone number.

        Returns the highest-scored record (prefers real data over
        'not found' / 'other operator' records).
        """
        msisdn = normalise_msisdn(number)
        recs = self._records.get(msisdn)
        if not recs:
            return None
        # Return best scored record
        return max(recs, key=_record_score)

    def lookup_all(self, number: str) -> List[SubscriberIdentification]:
        """Look up all identification records for a phone number.

        Returns all records sorted by score (best first).
        """
        msisdn = normalise_msisdn(number)
        recs = self._records.get(msisdn)
        if not recs:
            return []
        return sorted(recs, key=_record_score, reverse=True)

    def lookup_label(self, number: str) -> str:
        """Get a short display label for a phone number."""
        rec = self.lookup(number)
        if rec is None:
            return ""
        return rec.display_label

    def to_dict(self) -> Dict[str, Any]:
        """Export best record per MSISDN as a dict."""
        result = {}
        for msisdn, recs in self._records.items():
            best = max(recs, key=_record_score)
            result[msisdn] = best.to_dict()
        return result

    def get_all(self) -> List[SubscriberIdentification]:
        """Get all loaded records (flat list)."""
        result = []
        for recs in self._records.values():
            result.extend(recs)
        return result

    def get_all_best(self) -> List[SubscriberIdentification]:
        """Get best record per MSISDN."""
        return [max(recs, key=_record_score) for recs in self._records.values()]

    def get_pesel_groups(self) -> Dict[str, List[SubscriberIdentification]]:
        """Group all records by PESEL.

        Returns: {pesel: [records]} for PESELs that appear with data.
        Only includes records with actual PESEL values.
        """
        groups: Dict[str, List[SubscriberIdentification]] = {}
        for recs in self._records.values():
            for rec in recs:
                if rec.pesel and rec.has_owner_data:
                    if rec.pesel not in groups:
                        groups[rec.pesel] = []
                    groups[rec.pesel].append(rec)
        return groups

    def get_nip_groups(self) -> Dict[str, List[SubscriberIdentification]]:
        """Group all records by NIP (companies)."""
        groups: Dict[str, List[SubscriberIdentification]] = {}
        for recs in self._records.values():
            for rec in recs:
                if rec.nip and rec.has_owner_data:
                    if rec.nip not in groups:
                        groups[rec.nip] = []
                    groups[rec.nip].append(rec)
        return groups

    def enrich_records(self):
        """Cross-enrich records: fill missing data from other records of the same MSISDN.

        When one operator returns 'not found' but another has data,
        or when PESEL/address is missing in one record but present in another.
        """
        for msisdn, recs in self._records.items():
            if len(recs) < 2:
                continue
            # Find the best record to use as source
            best = max(recs, key=_record_score)
            if not best.has_owner_data:
                continue
            # Enrich other records
            for rec in recs:
                if rec is best:
                    continue
                if not rec.pesel and best.pesel:
                    rec.pesel = best.pesel
                if not rec.first_name and best.first_name:
                    rec.first_name = best.first_name
                if not rec.last_name and best.last_name:
                    rec.last_name = best.last_name
                if not rec.city and best.city:
                    rec.city = best.city
                if not rec.address and best.address:
                    rec.address = best.address
                if not rec.nip and best.nip:
                    rec.nip = best.nip
                if not rec.regon and best.regon:
                    rec.regon = best.regon

        # Cross-enrich by PESEL: if same PESEL appears on different MSISDNs,
        # fill missing fields from the most complete record
        pesel_groups = self.get_pesel_groups()
        for pesel, pesel_recs in pesel_groups.items():
            if len(pesel_recs) < 2:
                continue
            best = max(pesel_recs, key=_record_score)
            for rec in pesel_recs:
                if rec is best:
                    continue
                if not rec.first_name and best.first_name:
                    rec.first_name = best.first_name
                if not rec.last_name and best.last_name:
                    rec.last_name = best.last_name
                if not rec.city and best.city:
                    rec.city = best.city
                if not rec.address and best.address:
                    rec.address = best.address
                if not rec.document_number and best.document_number:
                    rec.document_number = best.document_number

    def clear(self):
        """Clear all loaded data."""
        self._records.clear()
        self._files_loaded.clear()
