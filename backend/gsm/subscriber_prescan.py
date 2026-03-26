"""Lightweight subscriber MSISDN pre-scan for GSM billing files.

Extracts the subscriber phone number from a billing file **without** doing
a full parse.  This is used during multi-file import to detect whether all
uploaded billings belong to the same person.  If they don't, the user is
asked to choose which subscriber to analyse.

Design goals:
  * Read at most 30–100 rows per file — fast even for large billings.
  * One function per operator format; each wrapped in try/except so a
    failure in one file never blocks the rest.
  * Forward-compatible with the future "Meeting" (multi-subscriber BTS
    cross-analysis) feature.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class FileSubscriberInfo:
    """Quick pre-scan result for one billing file."""

    filename: str
    path: Path
    operator_id: str  # "plus", "play", "tmobile", "orange", "orange_retencja", ""
    msisdn: str = ""  # normalised MSISDN or "" if undetectable
    confidence: float = 0.0  # 0.0 = unknown, 1.0 = certain
    detail: str = ""  # human-readable note, e.g. "Parametr column row 3"
    file_subtype: str = ""  # "pol", "td", "" — distinguishes complementary billing files

    def to_dict(self) -> Dict[str, Any]:
        return {
            "filename": self.filename,
            "operator_id": self.operator_id,
            "msisdn": self.msisdn,
            "confidence": round(self.confidence, 2),
            "detail": self.detail,
            "file_subtype": self.file_subtype,
        }


@dataclass
class SubscriberGrouping:
    """Result of grouping billing files by subscriber MSISDN."""

    is_single_subscriber: bool = True
    needs_confirmation: bool = False  # True when user should review file selection
    confirmation_reason: str = ""     # e.g. "complementary_files", "low_confidence", "multi_subscriber"
    subscribers: Dict[str, List[FileSubscriberInfo]] = field(default_factory=dict)
    undetected_files: List[FileSubscriberInfo] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "is_single_subscriber": self.is_single_subscriber,
            "needs_confirmation": self.needs_confirmation,
            "confirmation_reason": self.confirmation_reason,
            "subscribers": {
                msisdn: [f.to_dict() for f in files]
                for msisdn, files in self.subscribers.items()
            },
            "undetected_files": [f.to_dict() for f in self.undetected_files],
        }


# ---------------------------------------------------------------------------
# Phone number helpers  (mirrors logic from parsers/plus.py, base.py)
# ---------------------------------------------------------------------------

_PHONE_RE = re.compile(r"^\d{7,15}$")


def _normalize_phone(number: str) -> str:
    """Normalise a phone number (same logic as BillingParser.normalize_phone)."""
    if not number:
        return ""
    s = re.sub(r"[\s\-\(\)\.]+", "", number.strip())
    if s.startswith("00") and len(s) > 10:
        s = "+" + s[2:]
    if re.match(r"^\d{9}$", s):
        s = "+48" + s
    if re.match(r"^\d{10,}$", s) and not s.startswith("+"):
        s = "+" + s
    return s


def _digits(phone: str) -> str:
    """Return only digits from a phone string."""
    return re.sub(r"[^\d]", "", phone)


def _is_phone(text: str) -> bool:
    if not text:
        return False
    return bool(_PHONE_RE.match(re.sub(r"[\s\-\+\(\)\.]+", "", text)))


def is_same_subscriber(a: str, b: str) -> bool:
    """Compare two MSISDNs by last 9 digits (standard Polish number length)."""
    da, db = _digits(a), _digits(b)
    if not da or not db:
        return False
    if len(da) >= 9 and len(db) >= 9:
        return da[-9:] == db[-9:]
    return da == db


def _canonical(msisdn: str) -> str:
    """Return a canonical key for grouping (last 9 digits, +48 prefix)."""
    d = _digits(msisdn)
    if len(d) >= 9:
        return "+48" + d[-9:]
    return msisdn


# ---------------------------------------------------------------------------
# Per-operator quick extractors
# ---------------------------------------------------------------------------

def _detect_plus_subtype(path: Path) -> str:
    """Detect whether a Plus CSV is POL (calls/SMS) or TD (data sessions).

    Returns "pol", "td", or "" if unknown.
    """
    raw = path.read_bytes()
    for enc in ("cp1250", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        return ""

    first_line = text.split("\n", 1)[0].lower()
    if "uga/typ" in first_line:
        return "pol"
    if "czas trwania" in first_line:
        return "td"
    return ""


def _prescan_plus_pol(path: Path) -> Optional[str]:
    """Extract subscriber MSISDN from Plus POL CSV (Parametr column).

    Plus POL uses custom quoting and cp1250 encoding.
    Reads first ~25 rows after header.
    """
    raw = path.read_bytes()
    for enc in ("cp1250", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        return None

    lines = text.splitlines()[:50]

    # Find header row containing "Parametr" or "parametr"
    header_idx = -1
    parametr_col = -1
    for idx, line in enumerate(lines):
        parts = [p.strip().strip('"').lower() for p in line.split(",")]
        for ci, p in enumerate(parts):
            if "parametr" in p:
                header_idx = idx
                parametr_col = ci
                break
        if header_idx >= 0:
            break

    if header_idx < 0 or parametr_col < 0:
        return None

    # Scan data rows for a phone-like number in Parametr column
    for line in lines[header_idx + 1: header_idx + 20]:
        parts = [p.strip().strip('"') for p in line.split(",")]
        if parametr_col < len(parts):
            val = parts[parametr_col].strip()
            if re.match(r"^\d{6,15}$", val):
                return _normalize_phone(val)

    return None


def _prescan_plus_td(path: Path) -> Optional[str]:
    """Extract subscriber MSISDN from Plus TD CSV (MSISDN column).

    Plus TD has columns: Parametr, Start, Koniec, ..., MSISDN, IMEI, IMSI
    """
    raw = path.read_bytes()
    for enc in ("cp1250", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        return None

    lines = text.splitlines()[:50]

    header_idx = -1
    msisdn_col = -1
    parametr_col = -1
    for idx, line in enumerate(lines):
        parts = [p.strip().strip('"').lower() for p in line.split(",")]
        for ci, p in enumerate(parts):
            if p == "msisdn":
                msisdn_col = ci
            if "parametr" in p:
                parametr_col = ci
        if msisdn_col >= 0:
            header_idx = idx
            break

    if header_idx < 0:
        return None

    # Check first data rows for MSISDN value
    for line in lines[header_idx + 1: header_idx + 15]:
        parts = [p.strip().strip('"') for p in line.split(",")]
        if msisdn_col >= 0 and msisdn_col < len(parts):
            val = parts[msisdn_col].strip()
            if re.match(r"^\d{6,15}$", val):
                return _normalize_phone(val)
        # Fallback to Parametr column
        if parametr_col >= 0 and parametr_col < len(parts):
            val = parts[parametr_col].strip()
            if re.match(r"^\d{6,15}$", val):
                return _normalize_phone(val)

    return None


def _prescan_play(path: Path) -> Optional[str]:
    """Extract subscriber MSISDN from Play CSV by frequency analysis.

    Play uses semicolon-delimited CSV with cp1250 encoding.
    Subscriber number appears most frequently across UI_MSISDN / UW_MSISDN.
    """
    raw = path.read_bytes()
    for enc in ("cp1250", "utf-8", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        return None

    reader = csv.reader(io.StringIO(text), delimiter=";")
    header_idx = -1
    ui_col = -1
    uw_col = -1

    rows_read = 0
    for idx, row in enumerate(reader):
        if rows_read > 150:
            break
        rows_read += 1

        if header_idx < 0:
            lower = [c.strip().lower() for c in row]
            for ci, c in enumerate(lower):
                if c == "ui_msisdn":
                    ui_col = ci
                elif c == "uw_msisdn":
                    uw_col = ci
            if ui_col >= 0 or uw_col >= 0:
                header_idx = idx
            continue

        # Data row — count phone numbers
        break  # first pass: just find header

    if header_idx < 0:
        return None

    # Re-read for frequency counting
    reader2 = csv.reader(io.StringIO(text), delimiter=";")
    counter: Counter = Counter()
    for idx, row in enumerate(reader2):
        if idx <= header_idx:
            continue
        if idx > header_idx + 100:
            break
        if ui_col >= 0 and ui_col < len(row):
            val = row[ui_col].strip()
            if _is_phone(val):
                counter[_normalize_phone(val)] += 1
        if uw_col >= 0 and uw_col < len(row):
            val = row[uw_col].strip()
            if _is_phone(val):
                counter[_normalize_phone(val)] += 1

    if counter:
        return counter.most_common(1)[0][0]
    return None


def _prescan_xlsx_metadata(path: Path, max_rows: int = 30) -> Optional[str]:
    """Extract subscriber MSISDN from XLSX metadata header area.

    Works for T-Mobile and Orange — both store subscriber MSISDN in the
    metadata rows above the column headers.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active or wb.worksheets[0]
        for ri, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
            if ri >= max_rows:
                break
            text = " ".join(
                str(c).strip() for c in row if c is not None and str(c).strip()
            )
            if not text:
                continue
            # Pattern: "MSISDN: 48123456789" or "MSISDN:48123456789"
            m = re.search(
                r"(?:msisdn|numer\s*telefonu|numer\s*abonenta|nr\s*tel)"
                r"\s*:?\s*\+?(\d[\d\s\-]{7,})",
                text,
                re.I,
            )
            if m:
                candidate = re.sub(r"[\s\-]", "", m.group(1))
                if re.match(r"^\d{9,15}$", candidate):
                    return _normalize_phone(candidate)

            # Check individual cells for phone number patterns
            for cell in row:
                if cell is None:
                    continue
                cell_text = str(cell).strip()
                cleaned = cell_text.replace(" ", "").replace("-", "")
                if re.match(r"^\+?48?\d{9}$", cleaned):
                    return _normalize_phone(cell_text)
    finally:
        wb.close()

    return None


def _prescan_orange_retencja(path: Path) -> Optional[str]:
    """Extract subscriber MSISDN from Orange Retencja XLSX.

    Orange Retencja has metadata with MSISDN above data (similar to T-Mobile).
    Falls back to scanning first sheet metadata.
    """
    # Reuse the same XLSX metadata scanner
    return _prescan_xlsx_metadata(path, max_rows=30)


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def prescan_subscriber(
    filename: str,
    path: Path,
    operator_id: str,
) -> FileSubscriberInfo:
    """Quick pre-scan a billing file to extract subscriber MSISDN.

    Returns FileSubscriberInfo with msisdn="" and confidence=0.0 if
    the subscriber cannot be detected.
    """
    info = FileSubscriberInfo(
        filename=filename,
        path=path,
        operator_id=operator_id,
    )

    try:
        msisdn = None
        suffix = path.suffix.lower()

        if operator_id == "plus":
            # Detect subtype (POL=calls/SMS, TD=data) from header
            _subtype = _detect_plus_subtype(path)
            info.file_subtype = _subtype
            # Try POL first, then TD
            msisdn = _prescan_plus_pol(path)
            if msisdn:
                info.detail = f"Parametr column ({_subtype.upper() or 'POL'} format)"
            else:
                msisdn = _prescan_plus_td(path)
                if msisdn:
                    info.detail = f"MSISDN column ({_subtype.upper() or 'TD'} format)"

        elif operator_id == "play":
            msisdn = _prescan_play(path)
            if msisdn:
                info.detail = "UI/UW_MSISDN frequency analysis"

        elif operator_id == "tmobile":
            msisdn = _prescan_xlsx_metadata(path, max_rows=30)
            if msisdn:
                info.detail = "XLSX metadata header"

        elif operator_id == "orange":
            msisdn = _prescan_xlsx_metadata(path, max_rows=30)
            if msisdn:
                info.detail = "XLSX metadata header"

        elif operator_id == "orange_retencja":
            msisdn = _prescan_orange_retencja(path)
            if msisdn:
                info.detail = "XLSX metadata header"

        else:
            # Unknown operator — try XLSX metadata, then CSV Play-style
            if suffix in (".xlsx", ".xlsm", ".xls"):
                msisdn = _prescan_xlsx_metadata(path, max_rows=30)
                if msisdn:
                    info.detail = "XLSX metadata (generic)"
            elif suffix in (".csv", ".txt"):
                # Try Play-style frequency
                msisdn = _prescan_play(path)
                if msisdn:
                    info.detail = "CSV frequency analysis (generic)"

        if msisdn:
            info.msisdn = msisdn
            info.confidence = 0.9
        else:
            info.confidence = 0.0
            info.detail = "Nie udało się wykryć numeru abonenta"

    except Exception as exc:
        log.debug("Prescan failed for %s: %s", filename, exc)
        info.confidence = 0.0
        info.detail = f"Błąd prescan: {exc}"

    return info


# ---------------------------------------------------------------------------
# Grouping
# ---------------------------------------------------------------------------

def group_by_subscriber(
    billing_files: "List[Any]",  # List[ScannedFile] but avoid circular import
) -> SubscriberGrouping:
    """Pre-scan all billing files and group them by subscriber MSISDN.

    Parameters
    ----------
    billing_files : list
        List of ``ScannedFile`` objects (from folder_scanner) with attributes
        ``filename``, ``path``, ``operator_id``.

    Returns
    -------
    SubscriberGrouping
        ``is_single_subscriber`` is True when all detectable files belong to
        the same person (or no subscriber could be detected at all).
    """
    if len(billing_files) <= 1:
        return SubscriberGrouping(is_single_subscriber=True)

    infos: List[FileSubscriberInfo] = []
    for sf in billing_files:
        fsi = prescan_subscriber(
            filename=sf.filename,
            path=sf.path,
            operator_id=getattr(sf, "operator_id", "") or "",
        )
        infos.append(fsi)

    # Group by canonical MSISDN
    groups: Dict[str, List[FileSubscriberInfo]] = {}
    undetected: List[FileSubscriberInfo] = []

    for fsi in infos:
        if not fsi.msisdn:
            undetected.append(fsi)
            continue
        key = _canonical(fsi.msisdn)
        groups.setdefault(key, []).append(fsi)

    # Determine if single subscriber
    unique_subscribers = len(groups)
    is_single = unique_subscribers <= 1

    # Determine if confirmation is needed
    needs_confirm = False
    confirm_reason = ""

    if unique_subscribers > 1:
        # Multiple different subscribers — user must choose
        needs_confirm = True
        confirm_reason = "multi_subscriber"
    elif unique_subscribers == 1:
        files_in_group = list(groups.values())[0]
        if len(files_in_group) > 1:
            # Same subscriber, multiple files — complementary (e.g. POL + TD)
            # Ask user to confirm which files to include
            needs_confirm = True
            confirm_reason = "complementary_files"
        if undetected:
            # Some files couldn't be identified — uncertain
            needs_confirm = True
            confirm_reason = "undetected_files"
    elif unique_subscribers == 0 and len(undetected) > 1:
        # No subscriber detected in any file — very uncertain
        needs_confirm = True
        confirm_reason = "all_undetected"

    return SubscriberGrouping(
        is_single_subscriber=is_single,
        needs_confirmation=needs_confirm,
        confirmation_reason=confirm_reason,
        subscribers=groups,
        undetected_files=undetected,
    )
