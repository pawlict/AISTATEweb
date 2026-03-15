"""GSM billing processing pipeline.

Orchestrates: file upload → operator detection → parse → normalize → analyze → store.
Supports both XLSX and CSV billing formats.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import openpyxl

from .parsers.base import BillingParseResult, SubscriberInfo
from .parsers.registry import get_parser, detect_operator
from .normalize import normalize_records, extract_contact_numbers
from .subscriber import parse_subscriber_file

log = logging.getLogger(__name__)


def load_xlsx_sheets(
    path: Path,
    max_rows: int = 100_000,
    max_cols: int = 100,
) -> Dict[str, List[List[Any]]]:
    """Load all sheets from an XLSX file.

    Args:
        path: Path to the XLSX file.
        max_rows: Maximum rows per sheet (safety limit).
        max_cols: Maximum columns per sheet.

    Returns:
        Dict of sheet_name → rows (list of lists).
    """
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sheets: Dict[str, List[List[Any]]] = {}

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: List[List[Any]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    log.warning(
                        "Sheet '%s' truncated at %d rows", sheet_name, max_rows
                    )
                    break
                # Truncate columns
                cells = list(row[:max_cols])
                rows.append(cells)
            sheets[sheet_name] = rows
    finally:
        wb.close()

    return sheets


def _get_first_header(sheets: Dict[str, List[List[Any]]]) -> List[str]:
    """Extract first non-empty row (header candidate) from the first sheet."""
    for sheet_name, rows in sheets.items():
        for row in rows[:20]:
            cells = [str(c).strip().lower() for c in row if c is not None]
            non_empty = [c for c in cells if c]
            if len(non_empty) >= 3:
                return cells
    return []


def _load_billing_file(billing_path: Path) -> Dict[str, List[List[Any]]]:
    """Load billing file (XLSX or CSV) into sheets dict format.

    For XLSX: loads all sheets via openpyxl.
    For CSV: checks for Plus CSV (custom quoting), Play CSV (semicolon),
             or falls back to generic CSV loading.
    """
    suffix = billing_path.suffix.lower()

    if suffix == ".csv":
        # Check if it's a Plus CSV billing (custom quoting, comma-delimited)
        from .parsers.plus import is_plus_csv, load_plus_csv
        if is_plus_csv(billing_path):
            log.info("Detected Plus CSV billing: %s", billing_path.name)
            return load_plus_csv(billing_path)
        # Check if it's a Play CSV billing (semicolon-delimited)
        from .parsers.play import is_play_csv, load_play_csv
        if is_play_csv(billing_path):
            log.info("Detected Play CSV billing: %s", billing_path.name)
            return load_play_csv(billing_path)
        # Generic CSV fallback — load as single sheet
        import csv as csv_mod
        for encoding in ("cp1250", "utf-8-sig", "utf-8", "latin-1"):
            try:
                with open(billing_path, "r", encoding=encoding, errors="strict") as f:
                    reader = csv_mod.reader(f, delimiter=";")
                    rows: List[List[Any]] = []
                    for i, row in enumerate(reader):
                        if i >= 200_000:
                            break
                        rows.append(row)
                    if rows:
                        return {"CSV": rows}
            except (UnicodeDecodeError, UnicodeError):
                continue
        return {}

    # Default: XLSX
    return load_xlsx_sheets(billing_path)


def process_billing(
    billing_path: Path,
    subscriber_path: Optional[Path] = None,
    own_numbers: Optional[Set[str]] = None,
) -> BillingParseResult:
    """Full processing pipeline for a GSM billing file (XLSX or CSV).

    Steps:
    1. Load file (XLSX sheets or CSV rows)
    2. Auto-detect operator from headers/sheet names
    3. Parse billing records with operator-specific parser
    4. Optionally parse subscriber identification file
    5. Normalize records (phones, directions, dedup)
    6. Compute summary statistics

    Args:
        billing_path: Path to the billing XLSX/CSV file.
        subscriber_path: Optional path to subscriber identification XLSX.
        own_numbers: Optional set of own phone numbers for direction detection.

    Returns:
        BillingParseResult with all parsed and normalized data.
    """
    log.info("Processing billing: %s", billing_path.name)

    # 1. Load file (XLSX or CSV)
    sheets = _load_billing_file(billing_path)
    if not sheets:
        return BillingParseResult(
            warnings=["Nie udało się wczytać pliku (brak danych)"]
        )

    # 2. Detect operator
    headers = _get_first_header(sheets)
    sheet_names = [s.lower() for s in sheets.keys()]
    parser = get_parser(headers, sheet_names)

    log.info("Detected operator: %s (parser: %s)", parser.OPERATOR_NAME, parser.OPERATOR_ID)

    # 2b. Schema validation (pre-parse drift detection)
    try:
        from .parsers.schema_registry import SchemaRegistry
        from .parsers.schema_validator import SchemaValidator
        registry = SchemaRegistry()
        validator = SchemaValidator(registry)

        # Auto-bootstrap schemas if none exist
        if not registry.list_schemas():
            registry.bootstrap_all()

        validation = validator.validate(parser.OPERATOR_ID, headers)
        if validation.match_type in ("drift", "partial", "failed"):
            log.info(
                "Schema validation: %s (match=%s, confidence=%.2f, missing=%s)",
                parser.OPERATOR_ID, validation.match_type,
                validation.confidence, validation.missing_columns,
            )
    except Exception as exc:
        log.debug("Schema validation skipped: %s", exc)

    # 3. Parse billing
    result = parser.parse_workbook(sheets)

    # 4. Parse subscriber file if provided
    if subscriber_path and subscriber_path.exists():
        try:
            sub_sheets = load_xlsx_sheets(subscriber_path, max_rows=1000)
            for sheet_name, rows in sub_sheets.items():
                subscribers = parse_subscriber_file(rows, sheet_name)
                if subscribers:
                    # Merge subscriber info
                    sub = subscribers[0]
                    if sub.msisdn and not result.subscriber.msisdn:
                        result.subscriber.msisdn = sub.msisdn
                    if sub.owner_name and not result.subscriber.owner_name:
                        result.subscriber.owner_name = sub.owner_name
                    if sub.imsi and not result.subscriber.imsi:
                        result.subscriber.imsi = sub.imsi
                    if sub.imei and not result.subscriber.imei:
                        result.subscriber.imei = sub.imei
                    if sub.owner_address and not result.subscriber.owner_address:
                        result.subscriber.owner_address = sub.owner_address
                    if sub.owner_pesel and not result.subscriber.owner_pesel:
                        result.subscriber.owner_pesel = sub.owner_pesel
                    if sub.owner_id_number and not result.subscriber.owner_id_number:
                        result.subscriber.owner_id_number = sub.owner_id_number
                    if sub.activation_date and not result.subscriber.activation_date:
                        result.subscriber.activation_date = sub.activation_date
                    if sub.tariff and not result.subscriber.tariff:
                        result.subscriber.tariff = sub.tariff
                    if sub.sim_iccid and not result.subscriber.sim_iccid:
                        result.subscriber.sim_iccid = sub.sim_iccid
                    if sub.contract_type and not result.subscriber.contract_type:
                        result.subscriber.contract_type = sub.contract_type
                    # Store all subscribers in extra for multi-SIM scenarios
                    if len(subscribers) > 1:
                        result.subscriber.extra["additional_subscribers"] = [
                            s.to_dict() for s in subscribers[1:]
                        ]
                    break  # Use first sheet that has subscriber data
        except Exception as e:
            result.warnings.append(f"Błąd parsowania pliku abonenta: {e}")
            log.warning("Error parsing subscriber file: %s", e)

    # 5. Normalize
    effective_own = own_numbers or set()
    if result.subscriber.msisdn:
        effective_own.add(result.subscriber.msisdn)

    result = normalize_records(result, effective_own)

    log.info(
        "Parsed %d records for %s (%s)",
        len(result.records),
        result.subscriber.msisdn or "unknown",
        result.operator,
    )

    return result


def merge_billing_results(results: List[BillingParseResult]) -> BillingParseResult:
    """Merge multiple BillingParseResult objects into one.

    Used when complementary billing files (e.g. Plus POL for calls/SMS
    and Plus TD for data sessions) belong to the same subscriber and need
    to be analysed together.

    The first result with a non-empty subscriber serves as the base; records
    from all results are concatenated, and warnings are combined.
    """
    if not results:
        return BillingParseResult()
    if len(results) == 1:
        return results[0]

    # Pick the "primary" result — prefer the one with more records / subscriber info
    primary = max(results, key=lambda r: (
        1 if r.subscriber.msisdn else 0,
        len(r.records),
    ))

    merged = BillingParseResult(
        operator=primary.operator,
        operator_id=primary.operator_id,
        parser_version=primary.parser_version,
        subscriber=primary.subscriber,
        records=list(primary.records),
        summary=primary.summary,
        warnings=list(primary.warnings),
        sheet_name=primary.sheet_name,
        parse_method=primary.parse_method,
    )

    # Merge records from other results
    source_files = [primary.sheet_name or "primary"]
    for r in results:
        if r is primary:
            continue
        merged.records.extend(r.records)
        merged.warnings.extend(r.warnings)
        source_files.append(r.sheet_name or r.operator or "other")

        # Fill in missing subscriber fields from secondary results
        sub = r.subscriber
        if sub.msisdn and not merged.subscriber.msisdn:
            merged.subscriber.msisdn = sub.msisdn
        if sub.imsi and not merged.subscriber.imsi:
            merged.subscriber.imsi = sub.imsi
        if sub.imei and not merged.subscriber.imei:
            merged.subscriber.imei = sub.imei
        if sub.owner_name and not merged.subscriber.owner_name:
            merged.subscriber.owner_name = sub.owner_name

    # Recalculate summary
    from .parsers.base import BillingSummary
    s = BillingSummary()
    s.total_records = len(merged.records)
    dates = []
    for rec in merged.records:
        dur = rec.duration_seconds or 0
        cost = rec.cost or 0
        cost_g = rec.cost_gross or 0

        if rec.record_type == "CALL_OUT":
            s.calls_out += 1
            s.call_duration_seconds += dur
        elif rec.record_type == "CALL_IN":
            s.calls_in += 1
            s.call_duration_seconds += dur
        elif rec.record_type == "CALL_FORWARDED":
            s.calls_out += 1  # count in calls
            s.call_duration_seconds += dur
        elif rec.record_type == "SMS_OUT":
            s.sms_out += 1
        elif rec.record_type == "SMS_IN":
            s.sms_in += 1
        elif rec.record_type == "MMS_OUT":
            s.mms_out += 1
        elif rec.record_type == "MMS_IN":
            s.mms_in += 1
        elif rec.record_type == "DATA":
            s.data_sessions += 1

        s.total_duration_seconds += dur
        s.total_cost += cost
        s.total_cost_gross += cost_g

        if rec.date:
            dates.append(rec.date)

    if dates:
        s.period_from = min(dates)
        s.period_to = max(dates)

    contacts = set()
    for rec in merged.records:
        if rec.caller:
            contacts.add(rec.caller)
        if rec.callee:
            contacts.add(rec.callee)
    own = {merged.subscriber.msisdn} if merged.subscriber.msisdn else set()
    s.unique_contacts = len(contacts - own)
    merged.summary = s

    merged.warnings.insert(0, f"Scalono {len(results)} plików bilingowych: {', '.join(source_files)}")

    log.info("Merged %d billing results → %d total records", len(results), len(merged.records))

    return merged


def process_billing_batch(
    billing_paths: List[Path],
    subscriber_paths: Optional[List[Path]] = None,
) -> List[BillingParseResult]:
    """Process multiple billing files (e.g. multiple months or multiple subscribers).

    Args:
        billing_paths: List of billing XLSX file paths.
        subscriber_paths: Optional list of subscriber ID file paths
            (matched by index to billing_paths, or used as shared pool).

    Returns:
        List of BillingParseResult, one per billing file.
    """
    results: List[BillingParseResult] = []

    for i, bp in enumerate(billing_paths):
        sub_path = None
        if subscriber_paths:
            if i < len(subscriber_paths):
                sub_path = subscriber_paths[i]
            elif len(subscriber_paths) == 1:
                sub_path = subscriber_paths[0]  # shared subscriber file

        try:
            result = process_billing(bp, subscriber_path=sub_path)
            results.append(result)
        except Exception as e:
            log.error("Error processing %s: %s", bp.name, e)
            results.append(BillingParseResult(
                warnings=[f"Błąd przetwarzania {bp.name}: {e}"]
            ))

    return results
